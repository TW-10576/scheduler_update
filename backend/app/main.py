"""
FastAPI Main Application - Shift Scheduler V5.1
Complete with Employee Portal, Messaging, and Check-In/Out
"""

from fastapi import FastAPI, Depends, HTTPException, status, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from fastapi.responses import JSONResponse, StreamingResponse
import io
import calendar
from calendar import monthrange
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete, update, and_, or_, func, Float, Integer
from sqlalchemy.orm import selectinload, with_loader_criteria
from datetime import datetime, timedelta, date
from typing import List, Dict, Optional
from collections import defaultdict

from ortools.sat.python import cp_model

from app.config import settings
from app.database import get_db
from app.models import (
    User, Department, Manager, Employee, Role, Schedule, LeaveRequest,
    CheckInOut, Message, Notification,
    UserType, LeaveStatus, Attendance, Unavailability, Shift,
    OvertimeTracking, OvertimeRequest, OvertimeWorked, OvertimeStatus,
    LateNightWork, LeaveBalance, LeaveReminder, EmployeeWageConfig,
    PayrollCycle, WageCalculation, AttendanceSummary
)
from app.schemas import *
from app.auth import (
    get_password_hash, verify_password, create_access_token,
    get_current_active_user, require_admin, require_manager, require_employee
)
from app.schedule_generator import ShiftScheduleGenerator
from app.attendance_service import AttendanceService
from app.leave_reminder_service import LeaveReminderService
from app.wage_calculation_service import WageCalculationService

app = FastAPI(
    title="Shift Scheduler V5.1 API",
    description="Complete Employee Portal with Check-In/Out and Messaging",
    version="5.1.0"
)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=[str(origin) for origin in settings.CORS_ORIGINS],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# Custom exception handler for HTTPException to include CORS headers
@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    origin = request.headers.get("origin")
    cors_origins = [str(o) for o in settings.CORS_ORIGINS]
    
    response = JSONResponse(
        status_code=exc.status_code,
        content={"detail": exc.detail},
    )
    
    # Add CORS headers if origin is allowed
    if origin in cors_origins or "*" in cors_origins:
        response.headers["Access-Control-Allow-Origin"] = origin or "*"
        response.headers["Access-Control-Allow-Credentials"] = "true"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS, PATCH"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization, Accept"
    
    return response


# Exception handler for generic exceptions
@app.exception_handler(Exception)
async def general_exception_handler(request: Request, exc: Exception):
    print(f"Unhandled exception: {str(exc)}")
    import traceback
    traceback.print_exc()
    
    origin = request.headers.get("origin")
    cors_origins = [str(o) for o in settings.CORS_ORIGINS]
    
    response = JSONResponse(
        status_code=500,
        content={"detail": f"Internal server error: {str(exc)}"},
    )
    
    # Add CORS headers
    if origin in cors_origins or "*" in cors_origins:
        response.headers["Access-Control-Allow-Origin"] = origin or "*"
        response.headers["Access-Control-Allow-Credentials"] = "true"
        response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS, PATCH"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization, Accept"
    
    return response


# Health check
@app.get("/")
async def root():
    return {
        "message": "Shift Scheduler V5.1 API",
        "version": "5.1.0",
        "status": "running",
        "features": ["employee-portal", "messaging", "check-in-out", "leave-approval"]
    }


# Authentication
@app.post("/token", response_model=Token)
async def login(
    form_data: OAuth2PasswordRequestForm = Depends(),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(select(User).filter(User.username == form_data.username))
    user = result.scalar_one_or_none()
    
    if not user or not verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    # Update last login
    user.last_login = datetime.utcnow()
    await db.commit()
    
    # Get manager info if user is a manager
    manager_department_id = None
    if user.user_type == UserType.MANAGER:
        result = await db.execute(select(Manager).filter(Manager.user_id == user.id))
        manager = result.scalar_one_or_none()
        if manager:
            manager_department_id = manager.department_id
    
    access_token_expires = timedelta(minutes=settings.ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.username},
        expires_delta=access_token_expires
    )
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": {
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "full_name": user.full_name,
            "user_type": user.user_type,
            "manager_department_id": manager_department_id
        }
    }


@app.get("/users/me", response_model=UserResponse)
async def read_users_me(
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    # For managers, include their department_id
    if current_user.user_type == UserType.MANAGER:
        mgr_result = await db.execute(select(Manager).filter(Manager.user_id == current_user.id))
        manager = mgr_result.scalar_one_or_none()
        if manager:
            # Create a response with manager_department_id
            response_dict = {
                "id": current_user.id,
                "username": current_user.username,
                "email": current_user.email,
                "full_name": current_user.full_name,
                "user_type": current_user.user_type,
                "is_active": current_user.is_active,
                "manager_department_id": manager.department_id
            }
            return response_dict
    
    return current_user


# Admin: User Management
@app.post("/admin/users", response_model=UserResponse)
async def create_user(
    user_data: UserCreate,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    # Check if username exists
    result = await db.execute(select(User).filter(User.username == user_data.username))
    if result.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Username already exists")
    
    new_user = User(
        username=user_data.username,
        email=user_data.email,
        hashed_password=get_password_hash(user_data.password),
        full_name=user_data.full_name,
        user_type=user_data.user_type,
        is_active=True
    )
    
    db.add(new_user)
    await db.commit()
    await db.refresh(new_user)
    
    return new_user


@app.get("/admin/users", response_model=List[UserResponse])
async def list_users(
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(select(User))
    return result.scalars().all()


@app.delete("/admin/users/{user_id}")
async def delete_user(
    user_id: int,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """Delete a user (admin, manager, or employee)"""
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    result = await db.execute(select(User).filter(User.id == user_id))
    user = result.scalar_one_or_none()
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # If deleting a manager, delete the manager record
    if user.user_type == UserType.MANAGER:
        result = await db.execute(
            select(Manager).filter(Manager.user_id == user_id)
        )
        manager = result.scalar_one_or_none()
        if manager:
            await db.delete(manager)
    
    await db.delete(user)
    await db.commit()
    
    return {"message": f"User {user.username} deleted successfully"}


# Departments
@app.post("/departments", response_model=DepartmentResponse)
async def create_department(
    dept_data: DepartmentCreate,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    # Use provided dept_id or auto-generate if not provided
    dept_id_to_use = dept_data.dept_id.strip() if hasattr(dept_data, 'dept_id') and dept_data.dept_id else None
    
    if not dept_id_to_use:
        # Auto-generate dept_id (001, 002, 003, etc.)
        result = await db.execute(
            select(Department).order_by(Department.id.desc()).limit(1)
        )
        last_dept = result.scalar_one_or_none()
        next_id = 1
        if last_dept and last_dept.dept_id.isdigit():
            next_id = int(last_dept.dept_id) + 1
        dept_id_to_use = str(next_id).zfill(3)
    
    department = Department(
        dept_id=dept_id_to_use,
        name=dept_data.name,
        description=dept_data.description,
        is_active=True
    )
    
    db.add(department)
    await db.commit()
    await db.refresh(department)
    
    return department


@app.put("/departments/{department_id}", response_model=DepartmentResponse)
async def update_department(
    department_id: int,
    dept_data: DepartmentCreate,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(select(Department).filter(Department.id == department_id))
    department = result.scalar_one_or_none()
    
    if not department:
        raise HTTPException(status_code=404, detail="Department not found")
    
    department.name = dept_data.name
    department.description = dept_data.description
    department.updated_at = datetime.utcnow()
    
    await db.commit()
    await db.refresh(department)
    
    return department


@app.get("/departments", response_model=List[DepartmentResponse])
async def list_departments(
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    # All authenticated users can view active departments
    result = await db.execute(select(Department).filter(Department.is_active == True))
    return result.scalars().all()


@app.get("/departments/{department_id}/details", response_model=DepartmentDetailResponse)
async def get_department_details(
    department_id: int,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """Get department details with manager and employees + attendance info"""
    # Get department
    dept_result = await db.execute(
        select(Department).filter(Department.id == department_id)
    )
    department = dept_result.scalar_one_or_none()
    
    if not department:
        raise HTTPException(status_code=404, detail="Department not found")
    
    # Get manager for this department
    mgr_result = await db.execute(
        select(Manager).filter(Manager.department_id == department_id)
    )
    manager = mgr_result.scalar_one_or_none()
    manager_info = None
    if manager:
        user_result = await db.execute(
            select(User).filter(User.id == manager.user_id)
        )
        manager_user = user_result.scalar_one_or_none()
        if manager_user:
            manager_info = {
                "id": manager.id,
                "user_id": manager.user_id,
                "username": manager_user.username,
                "full_name": manager_user.full_name,
                "email": manager_user.email
            }
    
    # Get all employees in this department
    emp_result = await db.execute(
        select(Employee).filter(
            Employee.department_id == department_id,
            Employee.is_active == True
        ).order_by(Employee.id)
    )
    employees = emp_result.scalars().all()
    
    # Get today's date
    today = date.today()
    
    # Get latest attendance for each employee
    employee_list = []
    for emp in employees:
        # Get latest check-in/out
        checkin_result = await db.execute(
            select(CheckInOut).filter(
                CheckInOut.employee_id == emp.id
            ).order_by(CheckInOut.date.desc(), CheckInOut.check_in_time.desc()).limit(1)
        )
        latest_checkin = checkin_result.scalar_one_or_none()
        
        # Get today's schedule for assigned shift time
        schedule_result = await db.execute(
            select(Schedule).filter(
                Schedule.employee_id == emp.id,
                Schedule.date == today
            )
        )
        today_schedule = schedule_result.scalar_one_or_none()
        
        # Calculate total hours assigned
        total_hrs_assigned = None
        if today_schedule and today_schedule.start_time and today_schedule.end_time:
            start_h, start_m = map(int, today_schedule.start_time.split(':'))
            end_h, end_m = map(int, today_schedule.end_time.split(':'))
            start_decimal = start_h + start_m / 60
            end_decimal = end_h + end_m / 60
            total_hrs_assigned = end_decimal - start_decimal if end_decimal > start_decimal else 24 - start_decimal + end_decimal
        
        employee_list.append({
            "id": emp.id,
            "employee_id": emp.employee_id,
            "first_name": emp.first_name,
            "last_name": emp.last_name,
            "email": emp.email,
            "assigned_shift_time": f"{today_schedule.start_time} - {today_schedule.end_time}" if today_schedule and today_schedule.start_time and today_schedule.end_time else None,
            "total_hrs_assigned": f"{total_hrs_assigned:.2f}" if total_hrs_assigned else None,
            "latest_check_in": latest_checkin.check_in_time if latest_checkin else None,
            "latest_check_out": latest_checkin.check_out_time if latest_checkin else None
        })
    
    return {
        "id": department.id,
        "dept_id": department.dept_id,
        "name": department.name,
        "description": department.description,
        "manager": manager_info,
        "employees": employee_list
    }


@app.delete("/departments/{department_id}")
async def delete_department(
    department_id: int,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """Delete a department and all related data"""
    result = await db.execute(select(Department).filter(Department.id == department_id))
    department = result.scalar_one_or_none()
    
    if not department:
        raise HTTPException(status_code=404, detail="Department not found")
    
    dept_name = department.name
    
    # Delete all employees in the department (cascade will handle schedules, leave requests, check-ins)
    result = await db.execute(
        select(Employee).filter(Employee.department_id == department_id)
    )
    employees = result.scalars().all()
    for employee in employees:
        await db.delete(employee)
    
    # Delete all roles in the department (cascade will handle schedules)
    result = await db.execute(
        select(Role).filter(Role.department_id == department_id)
    )
    roles = result.scalars().all()
    for role in roles:
        await db.delete(role)
    
    # Delete all schedules in the department
    result = await db.execute(
        select(Schedule).filter(Schedule.department_id == department_id)
    )
    schedules = result.scalars().all()
    for schedule in schedules:
        await db.delete(schedule)
    
    # Delete the department itself
    await db.delete(department)
    await db.commit()
    
    return {"message": f"Department {dept_name} and all related data deleted successfully"}


@app.get("/departments/search/{query}")
async def search_departments(
    query: str,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """Search departments by ID, dept_id, or name"""
    # Try to search by internal ID first (if query is numeric)
    try:
        dept_id = int(query)
        result = await db.execute(
            select(Department).filter(
                Department.id == dept_id,
                Department.is_active == True
            )
        )
        department = result.scalar_one_or_none()
        if department:
            return {
                "id": department.id,
                "dept_id": department.dept_id,
                "name": department.name,
                "description": department.description
            }
    except ValueError:
        pass
    
    # Try to search by dept_id (3-digit code like 001, 002, etc.)
    result = await db.execute(
        select(Department).filter(
            Department.dept_id == query.zfill(3),
            Department.is_active == True
        )
    )
    department = result.scalar_one_or_none()
    if department:
        return {
            "id": department.id,
            "dept_id": department.dept_id,
            "name": department.name,
            "description": department.description
        }
    
    # Search by name (case-insensitive)
    result = await db.execute(
        select(Department).filter(
            Department.name.ilike(f"%{query}%"),
            Department.is_active == True
        )
    )
    departments = result.scalars().all()
    
    if departments:
        return [
            {
                "id": d.id,
                "dept_id": d.dept_id,
                "name": d.name,
                "description": d.description
            }
            for d in departments
        ]
    
    raise HTTPException(status_code=404, detail="Department not found")


# Managers
@app.post("/managers", response_model=dict)
async def create_manager(
    mgr_data: ManagerCreate,
    force_reassign: bool = False,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """Create a manager and link to a department. If force_reassign=true, will replace existing manager."""
    # Check if user exists
    result = await db.execute(select(User).filter(User.id == mgr_data.user_id))
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="User not found")
    
    # Check if department exists
    result = await db.execute(select(Department).filter(Department.id == mgr_data.department_id))
    if not result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Department not found")
    
    # Check if another manager is already assigned to this department
    existing_result = await db.execute(
        select(Manager).filter(
            and_(
                Manager.department_id == mgr_data.department_id,
                Manager.is_active == True
            )
        ).options(selectinload(Manager.user))
    )
    existing_manager = existing_result.scalar_one_or_none()
    
    if existing_manager:
        if not force_reassign:
            # Return existing manager info for frontend confirmation
            return {
                "status": "conflict",
                "message": "Manager already assigned to this department",
                "action_required": "reassign",
                "existing_manager": {
                    "id": existing_manager.id,
                    "user_id": existing_manager.user_id,
                    "username": existing_manager.user.username if existing_manager.user else None,
                    "full_name": existing_manager.user.full_name if existing_manager.user else None,
                    "email": existing_manager.user.email if existing_manager.user else None,
                    "department_id": existing_manager.department_id,
                    "is_active": existing_manager.is_active
                },
                "new_manager": {
                    "user_id": mgr_data.user_id,
                    "department_id": mgr_data.department_id
                }
            }
        else:
            # Reassign: unassign old manager from this department (don't deactivate)
            existing_manager.department_id = None
            existing_manager.updated_at = datetime.utcnow()
            db.add(existing_manager)  # Ensure the change is tracked
    
    # Create new manager
    manager = Manager(**mgr_data.dict())
    db.add(manager)
    await db.commit()
    await db.refresh(manager)
    
    return {
        "status": "success",
        "message": "Manager assigned successfully",
        "manager": {
            "id": manager.id,
            "user_id": manager.user_id,
            "department_id": manager.department_id,
            "is_active": manager.is_active
        }
    }


@app.get("/managers", response_model=List[ManagerDetailResponse])
async def list_managers(
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    from sqlalchemy.orm import selectinload
    result = await db.execute(
        select(Manager)
        .filter(Manager.is_active == True)
        .options(selectinload(Manager.user))
    )
    managers = result.scalars().all()
    
    # Build response with user details
    response = []
    for manager in managers:
        user = manager.user
        response.append({
            'id': manager.id,
            'user_id': manager.user_id,
            'username': user.username if user else None,
            'full_name': user.full_name if user else None,
            'email': user.email if user else None,
            'department_id': manager.department_id,
            'is_active': manager.is_active,
        })

    return response


@app.put("/managers/{manager_id}", response_model=ManagerResponse)
async def update_manager(
    manager_id: int,
    mgr_data: ManagerCreate,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(select(Manager).filter(Manager.id == manager_id))
    manager = result.scalar_one_or_none()
    
    if not manager:
        raise HTTPException(status_code=404, detail="Manager not found")
    
    # Check if another manager is already assigned to this department
    if mgr_data.department_id != manager.department_id:
        existing_result = await db.execute(
            select(Manager).filter(
                and_(
                    Manager.department_id == mgr_data.department_id,
                    Manager.is_active == True,
                    Manager.id != manager_id
                )
            )
        )
        existing_manager = existing_result.scalar_one_or_none()
        if existing_manager:
            raise HTTPException(
                status_code=409, 
                detail=f"Manager already assigned",
                headers={"X-Existing-Manager-Id": str(existing_manager.id)}
            )
    
    manager.department_id = mgr_data.department_id
    manager = mgr_data
    manager = mgr_data
    manager.updated_at = datetime.utcnow()
    
    await db.commit()
    await db.refresh(manager)
    
    return manager


@app.put("/managers/{manager_id}/reassign", response_model=ManagerResponse)
async def reassign_manager(
    manager_id: int,
    mgr_data: ManagerCreate,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """Reassign a manager to a department, removing any existing manager from that department"""
    result = await db.execute(select(Manager).filter(Manager.id == manager_id))
    manager = result.scalar_one_or_none()
    
    if not manager:
        raise HTTPException(status_code=404, detail="Manager not found")
    
    # If reassigning to a different department, unassign the existing manager
    if mgr_data.department_id != manager.department_id:
        existing_result = await db.execute(
            select(Manager).filter(
                and_(
                    Manager.department_id == mgr_data.department_id,
                    Manager.is_active == True,
                    Manager.id != manager_id
                )
            )
        )
        existing_manager = existing_result.scalar_one_or_none()
        if existing_manager:
            existing_manager.department_id = None
            existing_manager.updated_at = datetime.utcnow()
            db.add(existing_manager)  # Ensure the change is tracked
    
    manager.department_id = mgr_data.department_id
    manager = mgr_data
    manager = mgr_data
    manager.updated_at = datetime.utcnow()
    
    await db.commit()
    await db.refresh(manager)
    
    return manager


@app.delete("/managers/{manager_id}")
async def delete_manager(
    manager_id: int,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(select(Manager).filter(Manager.id == manager_id))
    manager = result.scalar_one_or_none()
    
    if not manager:
        raise HTTPException(status_code=404, detail="Manager not found")
    
    await db.delete(manager)
    await db.commit()
    
    return {"message": "Manager deleted successfully"}


# Employees
@app.post("/employees", response_model=EmployeeResponse)
async def create_employee(
    emp_data: EmployeeCreate,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    # Managers can only create in their department
    # Get the manager's department from Manager table
    result = await db.execute(select(Manager).filter(Manager.user_id == current_user.id))
    manager_record = result.scalar_one_or_none()
    
    if manager_record and emp_data.department_id != manager_record.department_id:
        raise HTTPException(status_code=403, detail="Can only create employees in your department")
    
    # Check if employee with this email already exists
    existing = await db.execute(select(Employee).filter(Employee.email == emp_data.email))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail=f"Employee with email {emp_data.email} already exists")

    # Generate numeric employee ID by extracting numbers from existing IDs
    result = await db.execute(select(Employee.employee_id))
    employee_ids = result.scalars().all()

    # Extract numeric parts from IDs like "EMP001", "EMP002", etc.
    max_num = 0
    for emp_id in employee_ids:
        if emp_id and len(emp_id) > 3:
            try:
                num = int(emp_id[3:])  # Extract numbers after "EMP"
                if num > max_num:
                    max_num = num
            except ValueError:
                pass

    new_employee_id = f"EMP{str(max_num + 1).zfill(3)}"  # Format as EMP001, EMP002, etc.

    # Create employee without user_id (optional)
    emp_dict = emp_data.dict(exclude={'password'})
    emp_dict['employee_id'] = new_employee_id
    employee = Employee(**emp_dict)
    db.add(employee)
    await db.flush()  # Get the employee ID
    
    # If password provided, create a user account
    if hasattr(emp_data, 'password') and emp_data.password:
        # Check if username already exists (use email as username)
        username = emp_data.email.split('@')[0]  # Use part before @ as username
        result = await db.execute(select(User).filter(User.username == username))
        if result.scalar_one_or_none():
            raise HTTPException(status_code=400, detail=f"Username {username} already exists")
        
        user = User(
            username=username,
            email=emp_data.email,
            full_name=f"{emp_data.first_name} {emp_data.last_name}",
            hashed_password=get_password_hash(emp_data.password),
            user_type=UserType.EMPLOYEE,
            is_active=True
        )
        db.add(user)
        await db.flush()
        
        # Link employee to user
        employee.user_id = user.id
    
    await db.commit()
    await db.refresh(employee)
    
    return employee


@app.get("/employees", response_model=List[EmployeeResponse])
async def list_employees(
    current_user: User = Depends(get_current_active_user),
    show_inactive: bool = False,  # Query parameter to show inactive employees
    db: AsyncSession = Depends(get_db)
):
    filters = []

    if current_user.user_type == UserType.ADMIN:
        # Admin sees all employees in their departments
        if not show_inactive:
            filters.append(Employee.is_active == True)
        result = await db.execute(select(Employee).filter(*filters) if filters else select(Employee))
    elif current_user.user_type == UserType.MANAGER:
        # Get manager's department from Manager table
        manager_result = await db.execute(select(Manager).filter(Manager.user_id == current_user.id))
        manager = manager_result.scalar_one_or_none()

        if manager:
            filters.append(Employee.department_id == manager.department_id)
            if not show_inactive:
                filters.append(Employee.is_active == True)
            result = await db.execute(select(Employee).filter(*filters))
        else:
            # No manager record, return empty list
            result = await db.execute(
                select(Employee).filter(Employee.id == -1)  # Returns empty
            )
    else:  # Employee
        if not show_inactive:
            filters.append(Employee.is_active == True)
        filters.append(Employee.user_id == current_user.id)
        result = await db.execute(select(Employee).filter(*filters))

    return result.scalars().all()


@app.put("/employees/{employee_id}", response_model=EmployeeResponse)
async def update_employee(
    employee_id: int,
    emp_data: EmployeeCreate,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(select(Employee).filter(Employee.id == employee_id))
    employee = result.scalar_one_or_none()
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    if current_user.user_type == UserType.MANAGER:
        # Get manager's department from Manager table
        manager_result = await db.execute(select(Manager).filter(Manager.user_id == current_user.id))
        manager = manager_result.scalar_one_or_none()
        
        if not manager or employee.department_id != manager.department_id:
            raise HTTPException(status_code=403, detail="Can only edit employees in your department")
    
    # Update employee fields (exclude password which needs special handling)
    emp_dict = emp_data.dict(exclude={'password'})
    for key, value in emp_dict.items():
        setattr(employee, key, value)
    
    # Handle password update if provided
    if emp_data.password:
        # Update user password if employee has a user account
        if employee.user_id:
            result = await db.execute(select(User).filter(User.id == employee.user_id))
            user = result.scalar_one_or_none()
            if user:
                user.hashed_password = get_password_hash(emp_data.password)
                # Ensure user is in the session
                db.add(user)
        else:
            # If no user account exists, create one
            username = emp_data.email.split('@')[0]
            result = await db.execute(select(User).filter(User.username == username))
            if not result.scalar_one_or_none():
                user = User(
                    username=username,
                    email=emp_data.email,
                    full_name=f"{emp_data.first_name} {emp_data.last_name}",
                    hashed_password=get_password_hash(emp_data.password),
                    user_type=UserType.EMPLOYEE,
                    is_active=True
                )
                db.add(user)
                await db.flush()
                employee.user_id = user.id
    
    employee.updated_at = datetime.utcnow()
    db.add(employee)
    await db.commit()
    await db.refresh(employee)
    
    return employee


@app.delete("/employees/{employee_id}")
async def delete_employee(
    employee_id: int,
    hard_delete: bool = False,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(select(Employee).filter(Employee.id == employee_id))
    employee = result.scalar_one_or_none()

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    if current_user.user_type == UserType.MANAGER:
        manager_dept = await get_manager_department(current_user, db)
        if not manager_dept or employee.department_id != manager_dept:
            raise HTTPException(status_code=403, detail="Can only delete employees in your department")

    # Get associated user if exists
    user = None
    if employee.user_id:
        user_result = await db.execute(select(User).filter(User.id == employee.user_id))
        user = user_result.scalar_one_or_none()

    if hard_delete:
        # Permanent deletion from database - delete both employee and user
        if user:
            await db.delete(user)
        await db.delete(employee)
        await db.commit()
        return {"message": "Employee and associated user permanently deleted"}
    else:
        # Soft delete - mark both as inactive
        employee.is_active = False
        if user:
            user.is_active = False
        await db.commit()
        return {"message": "Employee deleted successfully"}


# Roles
@app.post("/roles", response_model=RoleResponse)
async def create_role(
    role_data: RoleCreate,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    # Get manager's actual department assignment
    manager_dept = await get_manager_department(current_user, db)
    
    # Check if manager is trying to create role in a different department
    if current_user.user_type == UserType.MANAGER and role_data.department_id != manager_dept:
        raise HTTPException(status_code=403, detail="Can only create roles in your department")
    
    role = Role(**role_data.dict())
    db.add(role)
    await db.commit()
    await db.refresh(role)
    
    return role


@app.get("/roles", response_model=List[RoleDetailResponse])
async def list_roles(
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """List all roles with their shifts (eager loaded)"""
    stmt = select(Role).options(
        selectinload(Role.shifts),
        with_loader_criteria(Shift, Shift.is_active == True)
    )

    if current_user.user_type == UserType.ADMIN:
        stmt = stmt.filter(Role.is_active == True)
    else:
        # For managers, use get_manager_department helper
        manager_dept = await get_manager_department(current_user, db)
        stmt = stmt.filter(
            Role.department_id == manager_dept,
            Role.is_active == True
        )

    result = await db.execute(stmt)
    return result.scalars().unique().all()


@app.get("/roles/{role_id}", response_model=RoleDetailResponse)
async def get_role_detail(
    role_id: int,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Get role with all shifts"""
    stmt = select(Role).options(
        selectinload(Role.shifts),
        with_loader_criteria(Shift, Shift.is_active == True)
    ).filter(Role.id == role_id, Role.is_active == True)

    if current_user.user_type != UserType.ADMIN:
        manager_dept = await get_manager_department(current_user, db)
        stmt = stmt.filter(Role.department_id == manager_dept)

    result = await db.execute(stmt)
    role = result.scalar_one_or_none()
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    return role


@app.put("/roles/{role_id}", response_model=RoleResponse)
async def update_role(
    role_id: int,
    role_data: RoleUpdate,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    """Update a role"""
    result = await db.execute(
        select(Role).filter(Role.id == role_id)
    )
    role = result.scalar_one_or_none()

    if not role:
        raise HTTPException(status_code=404, detail="Role not found")

    # Get manager's actual department
    manager_dept = await get_manager_department(current_user, db)

    if role.department_id != manager_dept:
        raise HTTPException(status_code=403, detail="Can only update roles in your department")

    # Update only provided fields
    for key, value in role_data.dict(exclude_unset=True).items():
        if value is not None:
            setattr(role, key, value)

    await db.commit()
    await db.refresh(role)

    return role


@app.delete("/roles/{role_id}")
async def delete_role(
    role_id: int,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    """Delete a role (soft delete by marking inactive)"""
    result = await db.execute(
        select(Role).filter(Role.id == role_id)
    )
    role = result.scalar_one_or_none()
    
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    
    # Get manager's actual department
    manager_dept = await get_manager_department(current_user, db)
    
    if role.department_id != manager_dept:
        raise HTTPException(status_code=403, detail="Can only delete roles in your department")
    
    # Mark associated shifts as inactive so they disappear from role management views
    shift_result = await db.execute(
        select(Shift).filter(
            Shift.role_id == role_id,
            Shift.is_active == True
        )
    )
    role_shifts = shift_result.scalars().all()
    for shift in role_shifts:
        shift.is_active = False
    
    role.is_active = False
    await db.commit()
    
    return {"message": "Role and associated shifts deleted successfully"}

# Helper functions to resolve department ownership
async def get_user_department(user: User, db: AsyncSession) -> Optional[int]:
    """Resolve the department for a manager or employee user"""
    if user.user_type == UserType.MANAGER:
        result = await db.execute(select(Manager).filter(Manager.user_id == user.id))
        manager = result.scalar_one_or_none()
        return manager.department_id if manager else None
    
    if user.user_type == UserType.EMPLOYEE:
        result = await db.execute(select(Employee).filter(Employee.user_id == user.id))
        employee = result.scalar_one_or_none()
        return employee.department_id if employee else None
    
    return None


async def get_manager_department(user: User, db: AsyncSession) -> Optional[int]:
    """Get the department ID for a manager user"""
    if user.user_type != UserType.MANAGER:
        return None
    
    return await get_user_department(user, db)


# Employee: Check-In/Out
@app.post("/employee/check-in", response_model=CheckInResponse)
async def check_in(
    check_in_data: CheckInCreate,
    current_user: User = Depends(require_employee),
    db: AsyncSession = Depends(get_db)
):
    try:
        today = date.today()
        
        # Get employee by user_id
        result = await db.execute(
            select(Employee).filter(Employee.user_id == current_user.id)
        )
        employee = result.scalar_one_or_none()
        
        if not employee:
            raise HTTPException(status_code=400, detail="Employee record not found for this user")
        
        # Check if already checked in
        result = await db.execute(
            select(CheckInOut).filter(
                CheckInOut.employee_id == employee.id,
                CheckInOut.date == today,
                CheckInOut.check_out_time == None
            )
        )
        if result.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Already checked in today. Please check out first.")
        
        # Get today's schedule
        result = await db.execute(
            select(Schedule).filter(
                Schedule.employee_id == employee.id,
                Schedule.date == today
            )
        )
        schedule = result.scalar_one_or_none()
        
        if not schedule:
            raise HTTPException(status_code=400, detail="No scheduled shift for today. Please contact your manager.")
        
        # Calculate late status
        now = datetime.now()
        status_val = "on-time"
        
        try:
            if isinstance(schedule.start_time, str):
                scheduled_time = datetime.strptime(schedule.start_time, "%H:%M").time()
            else:
                scheduled_time = schedule.start_time if isinstance(schedule.start_time, type(None)) is False else datetime.strptime("09:00", "%H:%M").time()
            
            if scheduled_time:
                scheduled_datetime = datetime.combine(today, scheduled_time)
                diff_minutes = (now - scheduled_datetime).total_seconds() / 60
                
                if diff_minutes <= 0:
                    status_val = "on-time"
                elif diff_minutes <= 15:
                    status_val = "slightly-late"
                else:
                    status_val = "late"
        except (ValueError, TypeError, AttributeError) as e:
            # If we can't parse the time, just mark as on-time
            print(f"Time parsing error: {str(e)}")
            status_val = "on-time"
        
        # Create check-in record
        check_in = CheckInOut(
            employee_id=employee.id,
            schedule_id=schedule.id,
            date=today,
            check_in_time=now,
            check_in_status=status_val,
            location=check_in_data.location
        )
        
        db.add(check_in)
        await db.commit()
        
        # Refresh with proper eager loading
        result = await db.execute(
            select(CheckInOut)
            .where(CheckInOut.id == check_in.id)
            .options(
                selectinload(CheckInOut.employee),
                selectinload(CheckInOut.schedule).selectinload(Schedule.role)
            )
        )
        check_in = result.scalar_one()
        
        return check_in
    except HTTPException:
        raise
    except Exception as e:
        print(f"Check-in error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Check-in failed: {str(e)}")


@app.post("/employee/check-out", response_model=CheckInResponse)
async def check_out(
    check_out_data: CheckOutCreate,
    current_user: User = Depends(require_employee),
    db: AsyncSession = Depends(get_db)
):
    try:
        today = date.today()
        
        # Get employee by user_id
        result = await db.execute(
            select(Employee).filter(Employee.user_id == current_user.id)
        )
        employee = result.scalar_one_or_none()
        
        if not employee:
            raise HTTPException(status_code=400, detail="Employee record not found")

        # Find today's check-in
        result = await db.execute(
            select(CheckInOut).options(
                selectinload(CheckInOut.employee).selectinload(Employee.user),
                selectinload(CheckInOut.schedule).selectinload(Schedule.role)
            ).filter(
                CheckInOut.employee_id == employee.id,
                CheckInOut.date == today,
                CheckInOut.check_out_time == None
            )
        )
        check_in = result.scalar_one_or_none()

        if not check_in:
            raise HTTPException(status_code=400, detail="No active check-in found")

        check_in.check_out_time = datetime.now()
        check_in.notes = check_out_data.notes

        await db.commit()
        await db.refresh(check_in, ['employee', 'schedule'])

        # Create or update Attendance record with overtime calculation
        try:
            # Find existing attendance record
            att_result = await db.execute(
                select(Attendance).filter(
                    Attendance.employee_id == employee.id,
                    Attendance.date == today
                )
            )
            attendance = att_result.scalar_one_or_none()
            
            if not attendance:
                # Create new attendance record
                attendance = Attendance(
                    employee_id=employee.id,
                    schedule_id=check_in.schedule_id,
                    date=today,
                    in_time=check_in.check_in_time.strftime("%H:%M") if check_in.check_in_time else None,
                    out_time=check_in.check_out_time.strftime("%H:%M") if check_in.check_out_time else None,
                    status=check_in.check_in_status or "onTime"
                )
            else:
                # Update existing attendance record with checkout time
                attendance.in_time = check_in.check_in_time.strftime("%H:%M") if check_in.check_in_time else None
                attendance.out_time = check_in.check_out_time.strftime("%H:%M") if check_in.check_out_time else None
                attendance.schedule_id = check_in.schedule_id  # Ensure schedule_id is set
            
            # ==================== CHECK IF EMPLOYEE IS ON APPROVED LEAVE ====================
            # If employee has an approved leave for today, mark attendance status as "leave"
            from app.helpers import is_employee_on_approved_leave
            on_leave = await is_employee_on_approved_leave(db, employee.id, today)
            if on_leave:
                attendance.status = "leave"
            
            # Calculate worked hours and overtime
            if check_in.check_in_time and check_in.check_out_time:
                total_minutes = (check_in.check_out_time - check_in.check_in_time).total_seconds() / 60
                
                # Get break minutes from role, but only apply if total time is long enough
                break_minutes = 0
                if check_in.schedule and check_in.schedule.role:
                    role_break = check_in.schedule.role.break_minutes or 0
                    # Only apply break if total time is at least the break duration
                    if total_minutes >= role_break:
                        break_minutes = role_break
                
                worked_minutes = max(0, total_minutes - break_minutes)
                worked_hours = round(worked_minutes / 60, 2)
                
                attendance.worked_hours = worked_hours
                attendance.break_minutes = break_minutes
                
                # ==================== NIGHT WORK DETECTION ====================
                # Detect hours worked between 22:00 and 06:00 (late-night work)
                in_time = check_in.check_in_time.time()
                out_time = check_in.check_out_time.time()
                night_hours = await AttendanceService.calculate_night_work_hours(
                    in_time.strftime("%H:%M"),
                    out_time.strftime("%H:%M")
                )
                attendance.night_hours = night_hours
                
                # Calculate night allowance if configured
                if night_hours > 0 and employee.wage_config:
                    night_allowance = night_hours * employee.wage_config.hourly_rate * employee.wage_config.night_shift_multiplier
                    attendance.night_allowance = round(night_allowance, 2)
                
                # Store in LateNightWork table if there are night hours
                if night_hours > 0:
                    night_record_result = await db.execute(
                        select(LateNightWork).where(
                            and_(
                                LateNightWork.employee_id == employee.id,
                                LateNightWork.work_date == today
                            )
                        )
                    )
                    night_record = night_record_result.scalar_one_or_none()
                    
                    if not night_record:
                        night_record = LateNightWork(
                            employee_id=employee.id,
                            work_date=today,
                            night_start_time=in_time.strftime("%H:%M"),
                            night_end_time=out_time.strftime("%H:%M"),
                            night_hours=night_hours,
                            night_allowance_rate=employee.wage_config.night_shift_multiplier if employee.wage_config else 1.5,
                            night_allowance_amount=attendance.night_allowance
                        )
                        db.add(night_record)
                    else:
                        night_record.night_hours = night_hours
                        night_record.night_allowance_rate = employee.wage_config.night_shift_multiplier if employee.wage_config else 1.5
                        night_record.night_allowance_amount = attendance.night_allowance
                        db.add(night_record)
                
                # ==================== OVERTIME CALCULATION ====================
                # Rule: Calculate overtime based on approved overtime window
                # If 1hr approved OT 18-19 and shift is 9-18:
                #   - Checkout at 19:00 → 1hr OT
                #   - Checkout at 19:30 → 1hr OT (capped by approval)
                #   - Checkout at 18:30 → 0.5hr OT (proportional)
                
                attendance.overtime_hours = 0.0
                
                # Get approved overtime for this date if exists
                approved_ot_result = await db.execute(
                    select(OvertimeRequest).filter(
                        OvertimeRequest.employee_id == employee.id,
                        OvertimeRequest.request_date == today,
                        OvertimeRequest.status == OvertimeStatus.APPROVED
                    )
                )
                overtime_request = approved_ot_result.scalar_one_or_none()
                
                if overtime_request and check_in.schedule:
                    # Parse shift end time
                    try:
                        shift_end_str = check_in.schedule.end_time
                        if isinstance(shift_end_str, str):
                            shift_end_h, shift_end_m = map(int, shift_end_str.split(':'))
                        else:
                            shift_end_h, shift_end_m = shift_end_str.hour, shift_end_str.minute
                        
                        shift_end_time = datetime.combine(today, datetime.min.time().replace(hour=shift_end_h, minute=shift_end_m))
                        
                        # Parse approved OT window (from_time to to_time)
                        ot_start_str = overtime_request.from_time
                        ot_end_str = overtime_request.to_time
                        
                        if ot_start_str and ot_end_str:
                            ot_start_h, ot_start_m = map(int, ot_start_str.split(':'))
                            ot_end_h, ot_end_m = map(int, ot_end_str.split(':'))
                            
                            ot_window_start = datetime.combine(today, datetime.min.time().replace(hour=ot_start_h, minute=ot_start_m))
                            ot_window_end = datetime.combine(today, datetime.min.time().replace(hour=ot_end_h, minute=ot_end_m))
                            
                            # Calculate how much OT was actually worked
                            # OT period: from shift end to checkout time, capped by approval window
                            # Example: Shift ends 18:00, approved OT 18:00-20:00, checkout 19:30
                            # → OT worked: 18:00-19:30 = 1.5 hours (within 2-hour window)
                            
                            ot_period_start = shift_end_time  # When OT begins (shift end)
                            ot_period_end = check_in.check_out_time  # When employee actually left
                            
                            # Cap the actual OT by the approved window
                            actual_ot_start = max(ot_period_start, ot_window_start)
                            actual_ot_end = min(ot_period_end, ot_window_end)
                            
                            if actual_ot_end > actual_ot_start:
                                actual_ot_minutes = (actual_ot_end - actual_ot_start).total_seconds() / 60
                                actual_ot_hours = actual_ot_minutes / 60
                                
                                # Cap at approved hours
                                overtime_hours = min(actual_ot_hours, overtime_request.request_hours)
                                attendance.overtime_hours = round(overtime_hours, 2)
                        else:
                            # No specific time window, use approved hours if worked > 8hrs/day
                            actual_overtime = worked_hours - employee.daily_max_hours
                            if actual_overtime > 0:
                                overtime_hours = min(actual_overtime, overtime_request.request_hours)
                                attendance.overtime_hours = round(overtime_hours, 2)
                    except Exception as e:
                        print(f"Error parsing OT times: {str(e)}")
                        # Fallback: simple calculation
                        actual_overtime = worked_hours - employee.daily_max_hours
                        if actual_overtime > 0:
                            overtime_hours = min(actual_overtime, overtime_request.request_hours)
                            attendance.overtime_hours = round(overtime_hours, 2)
                elif worked_hours > employee.daily_max_hours:
                    # No approved OT, but worked more than 8hrs - show actual OT
                    actual_overtime = worked_hours - employee.daily_max_hours
                    attendance.overtime_hours = round(actual_overtime, 2)
                # ==================== END OVERTIME CALCULATION ====================
            
            db.add(attendance)
            await db.commit()
        except Exception as e:
            print(f"Error creating attendance record: {str(e)}")
            import traceback
            traceback.print_exc()
            # Don't fail the checkout, just log the error
        
        return check_in
    except HTTPException:
        raise
    except Exception as e:
        print(f"Check-out error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Check-out failed: {str(e)}")


@app.post("/attendance/record")
async def record_attendance(
    attendance_data: dict,
    current_user: User = Depends(require_employee),
    db: AsyncSession = Depends(get_db)
):
    """Record attendance for a schedule (check-in)"""
    try:
        today = date.today()
        
        # Get employee by user_id
        result = await db.execute(
            select(Employee).filter(Employee.user_id == current_user.id)
        )
        employee = result.scalar_one_or_none()
        
        if not employee:
            raise HTTPException(status_code=400, detail="Employee record not found")
        
        # Get schedule
        schedule_id = attendance_data.get('schedule_id')
        schedule_result = await db.execute(
            select(Schedule).filter(Schedule.id == schedule_id)
        )
        schedule = schedule_result.scalar_one_or_none()
        
        if not schedule:
            raise HTTPException(status_code=400, detail="Schedule not found")
        
        # Create attendance record
        attendance = Attendance(
            employee_id=employee.id,
            schedule_id=schedule_id,
            date=today,
            in_time=attendance_data.get('in_time'),
            status=attendance_data.get('status', 'onTime'),
            break_minutes=schedule.role.break_minutes if schedule.role else 0
        )
        
        db.add(attendance)
        await db.commit()
        await db.refresh(attendance)
        
        return {
            "id": attendance.id,
            "employee_id": attendance.employee_id,
            "schedule_id": attendance.schedule_id,
            "date": attendance.date,
            "in_time": attendance.in_time,
            "out_time": attendance.out_time,
            "status": attendance.status,
            "worked_hours": attendance.worked_hours,
            "overtime_hours": attendance.overtime_hours,
            "message": "Attendance recorded successfully"
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Attendance record error: {str(e)}")
        import traceback
        traceback.print_exc()


# Attendance Management
@app.get("/attendance")
async def get_attendance(
    start_date: date = None,
    end_date: date = None,
    employee_id: int = None,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Get attendance records with optional filters"""
    query = select(Attendance).options(
        selectinload(Attendance.employee),
        selectinload(Attendance.schedule).selectinload(Schedule.role)
    )

    # Role-based filtering
    target_employee = None
    manager_dept = None
    if current_user.user_type == UserType.EMPLOYEE:
        # Get employee by user_id
        emp_result = await db.execute(
            select(Employee).filter(Employee.user_id == current_user.id)
        )
        target_employee = emp_result.scalar_one_or_none()
        if target_employee:
            query = query.filter(Attendance.employee_id == target_employee.id)
        else:
            return []
    elif current_user.user_type == UserType.MANAGER:
        # Get manager's actual department
        manager_dept = await get_manager_department(current_user, db)
        if manager_dept:
            subquery = select(Employee.id).filter(Employee.department_id == manager_dept)
            query = query.filter(Attendance.employee_id.in_(subquery))
        else:
            return []

    # Additional filters
    if employee_id and current_user.user_type != UserType.EMPLOYEE:
        query = query.filter(Attendance.employee_id == employee_id)

    if start_date:
        query = query.filter(Attendance.date >= start_date)
    if end_date:
        query = query.filter(Attendance.date <= end_date)

    result = await db.execute(query.order_by(Attendance.date.desc()))
    attendance_records = list(result.scalars().all())
    
    # For today, also check for active check-ins (CheckInOut without checkout)
    today = date.today()
    if not start_date or start_date <= today:
        if not end_date or end_date >= today:
            # Get active check-ins for today
            if current_user.user_type == UserType.EMPLOYEE and target_employee:
                checkin_query = select(CheckInOut).filter(
                    CheckInOut.employee_id == target_employee.id,
                    CheckInOut.date == today,
                    CheckInOut.check_out_time == None
                ).options(
                    selectinload(CheckInOut.employee),
                    selectinload(CheckInOut.schedule).selectinload(Schedule.role)
                )
            elif current_user.user_type == UserType.MANAGER and manager_dept:
                subquery = select(Employee.id).filter(Employee.department_id == manager_dept)
                checkin_query = select(CheckInOut).filter(
                    CheckInOut.employee_id.in_(subquery),
                    CheckInOut.date == today,
                    CheckInOut.check_out_time == None
                ).options(
                    selectinload(CheckInOut.employee),
                    selectinload(CheckInOut.schedule).selectinload(Schedule.role)
                )
            else:
                checkin_query = select(CheckInOut).filter(
                    CheckInOut.date == today,
                    CheckInOut.check_out_time == None
                ).options(
                    selectinload(CheckInOut.employee),
                    selectinload(CheckInOut.schedule).selectinload(Schedule.role)
                )
            
            checkin_result = await db.execute(checkin_query)
            active_checkins = list(checkin_result.scalars().all())
            
            # Convert CheckInOut to Attendance-like response for return
            for checkin in active_checkins:
                # Check if there's already an Attendance record for this employee today
                existing = any(r.employee_id == checkin.employee_id and r.date == today for r in attendance_records)
                if not existing:
                    # Convert CheckInOut to dict matching AttendanceResponse
                    att_dict = {
                        'id': checkin.id,
                        'employee_id': checkin.employee_id,
                        'schedule_id': checkin.schedule_id,
                        'date': checkin.date,
                        'in_time': checkin.check_in_time.strftime("%H:%M") if checkin.check_in_time else None,
                        'out_time': None,
                        'status': checkin.check_in_status or 'onTime',
                        'out_status': None,
                        'worked_hours': 0.0,
                        'overtime_hours': 0.0,
                        'break_minutes': 0,
                        'notes': None,
                        'created_at': checkin.date,
                        'employee': checkin.employee,
                        'schedule': checkin.schedule
                    }
                    attendance_records.append(att_dict)
    
    return attendance_records


@app.get("/attendance/today")
async def get_todays_attendance(
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    """Get today's attendance summary for manager"""
    today = date.today()

    # Get all employees in department
    manager_dept = await get_manager_department(current_user, db)
    if not manager_dept:
        raise HTTPException(status_code=400, detail="Manager department not found")
    
    emp_result = await db.execute(
        select(Employee).filter(
            Employee.department_id == manager_dept,
            Employee.is_active == True
        )
    )
    employees = emp_result.scalars().all()

    # Get today's schedules
    sched_result = await db.execute(
        select(Schedule).filter(
            Schedule.department_id == manager_dept,
            Schedule.date == today
        )
    )
    schedules = sched_result.scalars().all()

    # Get today's check-ins
    checkin_result = await db.execute(
        select(CheckInOut).filter(
            CheckInOut.date == today
        )
    )
    checkins = {c.employee_id: c for c in checkin_result.scalars().all()}

    # Build attendance summary
    attendance_summary = []
    for schedule in schedules:
        checkin = checkins.get(schedule.employee_id)

        status = "absent"
        if checkin:
            if checkin.check_out_time:
                status = "completed"
            elif checkin.check_in_status == "on-time":
                status = "present"
            else:
                status = checkin.check_in_status

        attendance_summary.append({
            "employee_id": schedule.employee_id,
            "schedule_id": schedule.id,
            "scheduled_time": f"{schedule.start_time} - {schedule.end_time}",
            "check_in_time": checkin.check_in_time.isoformat() if checkin and checkin.check_in_time else None,
            "check_out_time": checkin.check_out_time.isoformat() if checkin and checkin.check_out_time else None,
            "status": status
        })

    return {
        "date": today.isoformat(),
        "total_scheduled": len(schedules),
        "total_checked_in": len(checkins),
        "attendance": attendance_summary
    }


@app.get("/attendance/stats")
async def get_attendance_stats(
    start_date: date,
    end_date: date,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Get attendance statistics"""
    query = select(CheckInOut).filter(
        CheckInOut.date >= start_date,
        CheckInOut.date <= end_date
    )

    if current_user.user_type == UserType.EMPLOYEE:
        query = query.filter(CheckInOut.employee_id == current_user.employee_id)
    elif current_user.user_type == UserType.MANAGER:
        manager_dept = await get_manager_department(current_user, db)
        if manager_dept:
            subquery = select(Employee.id).filter(Employee.department_id == manager_dept)
            query = query.filter(CheckInOut.employee_id.in_(subquery))
        else:
            query = query.filter(CheckInOut.employee_id == -1)  # Return empty

    result = await db.execute(query)
    records = result.scalars().all()

    total = len(records)
    on_time = len([r for r in records if r.check_in_status == "on-time"])
    late = len([r for r in records if r.check_in_status in ["slightly-late", "late"]])

    return {
        "total_days": total,
        "on_time": on_time,
        "late": late,
        "on_time_percentage": round((on_time / total * 100) if total > 0 else 0, 2)
    }


# Attendance Reports (Excel Export)
@app.get("/attendance/export/monthly")
async def export_monthly_attendance(
    department_id: int,
    year: int,
    month: int,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Export monthly attendance report as Excel"""
    try:
        # Check authorization: only admins can download any department, managers only their own
        if current_user.user_type == UserType.MANAGER:
            # For managers, check if they manage this department
            manager_result = await db.execute(
                select(Manager).filter(Manager.user_id == current_user.id, Manager.department_id == department_id)
            )
            if not manager_result.scalar_one_or_none():
                raise HTTPException(status_code=403, detail="You don't have permission to download reports for this department")
        elif current_user.user_type != UserType.ADMIN:
            raise HTTPException(status_code=403, detail="Only admins and managers can download attendance reports")
        
        # Get department
        dept_result = await db.execute(select(Department).filter(Department.id == department_id))
        department = dept_result.scalar_one_or_none()
        if not department:
            raise HTTPException(status_code=404, detail="Department not found")
        
        # Get employees in department
        emp_result = await db.execute(
            select(Employee).filter(Employee.department_id == department_id, Employee.is_active == True)
        )
        employees = emp_result.scalars().all()
        
        # Get attendance for the month
        start_date = date(year, month, 1)
        end_date = date(year, month, monthrange(year, month)[1])
        
        att_result = await db.execute(
            select(Attendance).filter(
                Attendance.employee_id.in_([e.id for e in employees]) if employees else False,
                Attendance.date >= start_date,
                Attendance.date <= end_date
            ).order_by(Attendance.date, Attendance.employee_id)
        )
        attendance_records = att_result.scalars().all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance"
        
        # Header styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = f"{department.name} - Monthly Attendance Report"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:J1')
        
        ws['A2'] = f"December 2025"
        ws.merge_cells('A2:J2')
        
        # Headers
        headers = ['Employee ID', 'Name', 'Date', 'Assigned Shift', 'Total Hrs Assigned', 'Check-In', 'Check-Out', 'Total Hrs Worked', 'Break Time', 'Overtime Hours', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Get schedules for shift information
        sched_result = await db.execute(
            select(Schedule).filter(
                Schedule.employee_id.in_([e.id for e in employees]) if employees else False,
                Schedule.date >= start_date,
                Schedule.date <= end_date
            )
        )
        schedules = sched_result.scalars().all()
        schedule_map = {}
        for sched in schedules:
            schedule_map[(sched.employee_id, sched.date)] = sched
        
        # Data
        row = 5
        for record in attendance_records:
            employee = next((e for e in employees if e.id == record.employee_id), None)
            if employee:
                schedule = schedule_map.get((record.employee_id, record.date))
                total_hrs_assigned = '-'
                assigned_shift = '-'
                if schedule and schedule.start_time and schedule.end_time:
                    assigned_shift = f"{schedule.start_time} - {schedule.end_time}"
                    try:
                        start_h, start_m = map(int, schedule.start_time.split(':'))
                        end_h, end_m = map(int, schedule.end_time.split(':'))
                        start_decimal = start_h + start_m / 60
                        end_decimal = end_h + end_m / 60
                        hours = end_decimal - start_decimal if end_decimal > start_decimal else 24 - start_decimal + end_decimal
                        total_hrs_assigned = f"{hours:.2f}"
                    except:
                        pass
                
                ws.cell(row=row, column=1).value = employee.employee_id
                ws.cell(row=row, column=2).value = f"{employee.first_name} {employee.last_name}"
                ws.cell(row=row, column=3).value = record.date.isoformat()
                ws.cell(row=row, column=4).value = assigned_shift
                ws.cell(row=row, column=5).value = total_hrs_assigned
                ws.cell(row=row, column=6).value = record.in_time or '-'
                ws.cell(row=row, column=7).value = record.out_time or '-'
                ws.cell(row=row, column=8).value = f"{record.worked_hours:.2f}" if record.worked_hours else '-'
                ws.cell(row=row, column=9).value = f"{record.break_minutes}" if record.break_minutes else '-'
                ws.cell(row=row, column=10).value = f"{record.overtime_hours:.2f}" if record.overtime_hours else '-'
                ws.cell(row=row, column=11).value = record.status or '-'
                for col in range(1, 12):
                    ws.cell(row=row, column=col).border = border
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 12
        
        # Save to bytes
        file_bytes = io.BytesIO()
        wb.save(file_bytes)
        file_bytes.seek(0)
        
        return StreamingResponse(
            iter([file_bytes.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={department.name}_attendance_{year}-{month:02d}.xlsx"}
        )
    except Exception as e:
        print(f"Export error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@app.get("/attendance/export/weekly")
async def export_weekly_attendance(
    department_id: int,
    start_date: date,
    end_date: date,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Export weekly attendance report as Excel"""
    try:
        # Check authorization: only admins can download any department, managers only their own
        if current_user.user_type == UserType.MANAGER:
            # For managers, check if they manage this department
            manager_result = await db.execute(
                select(Manager).filter(Manager.user_id == current_user.id, Manager.department_id == department_id)
            )
            if not manager_result.scalar_one_or_none():
                raise HTTPException(status_code=403, detail="You don't have permission to download reports for this department")
        elif current_user.user_type != UserType.ADMIN:
            raise HTTPException(status_code=403, detail="Only admins and managers can download attendance reports")
        
        # Get department
        dept_result = await db.execute(select(Department).filter(Department.id == department_id))
        department = dept_result.scalar_one_or_none()
        if not department:
            raise HTTPException(status_code=404, detail="Department not found")
        
        # Get employees in department
        emp_result = await db.execute(
            select(Employee).filter(Employee.department_id == department_id, Employee.is_active == True)
        )
        employees = emp_result.scalars().all()
        
        # Get attendance for the week
        att_result = await db.execute(
            select(Attendance).filter(
                Attendance.employee_id.in_([e.id for e in employees]) if employees else False,
                Attendance.date >= start_date,
                Attendance.date <= end_date
            ).order_by(Attendance.date, Attendance.employee_id)
        )
        attendance_records = att_result.scalars().all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Weekly"
        
        # Header styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = f"{department.name} - Weekly Attendance Report"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:J1')
        
        ws['A2'] = f"{start_date.isoformat()} to {end_date.isoformat()}"
        ws.merge_cells('A2:J2')
        
        # Headers
        headers = ['Employee ID', 'Name', 'Date', 'Assigned Shift', 'Total Hrs Assigned', 'Check-In', 'Check-Out', 'Total Hrs Worked', 'Break Time', 'Overtime Hours', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Get schedules for shift information
        sched_result = await db.execute(
            select(Schedule).filter(
                Schedule.employee_id.in_([e.id for e in employees]) if employees else False,
                Schedule.date >= start_date,
                Schedule.date <= end_date
            )
        )
        schedules = sched_result.scalars().all()
        schedule_map = {}
        for sched in schedules:
            schedule_map[(sched.employee_id, sched.date)] = sched
        
        # Data
        row = 5
        for record in attendance_records:
            employee = next((e for e in employees if e.id == record.employee_id), None)
            if employee:
                schedule = schedule_map.get((record.employee_id, record.date))
                total_hrs_assigned = '-'
                assigned_shift = '-'
                if schedule and schedule.start_time and schedule.end_time:
                    assigned_shift = f"{schedule.start_time} - {schedule.end_time}"
                    try:
                        start_h, start_m = map(int, schedule.start_time.split(':'))
                        end_h, end_m = map(int, schedule.end_time.split(':'))
                        start_decimal = start_h + start_m / 60
                        end_decimal = end_h + end_m / 60
                        hours = end_decimal - start_decimal if end_decimal > start_decimal else 24 - start_decimal + end_decimal
                        total_hrs_assigned = f"{hours:.2f}"
                    except:
                        pass
                
                ws.cell(row=row, column=1).value = employee.employee_id
                ws.cell(row=row, column=2).value = f"{employee.first_name} {employee.last_name}"
                ws.cell(row=row, column=3).value = record.date.isoformat()
                ws.cell(row=row, column=4).value = assigned_shift
                ws.cell(row=row, column=5).value = total_hrs_assigned
                ws.cell(row=row, column=6).value = record.in_time or '-'
                ws.cell(row=row, column=7).value = record.out_time or '-'
                ws.cell(row=row, column=8).value = f"{record.worked_hours:.2f}" if record.worked_hours else '-'
                ws.cell(row=row, column=9).value = f"{record.break_minutes}" if record.break_minutes else '-'
                ws.cell(row=row, column=10).value = f"{record.overtime_hours:.2f}" if record.overtime_hours else '-'
                ws.cell(row=row, column=11).value = record.status or '-'
                for col in range(1, 12):
                    ws.cell(row=row, column=col).border = border
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 12
        
        # Save to bytes
        file_bytes = io.BytesIO()
        wb.save(file_bytes)
        file_bytes.seek(0)
        
        return StreamingResponse(
            iter([file_bytes.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={department.name}_attendance_{start_date.isoformat()}_to_{end_date.isoformat()}.xlsx"}
        )
    except Exception as e:
        print(f"Weekly export error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@app.get("/attendance/export/employee-monthly")
async def export_employee_monthly_attendance(
    year: int,
    month: int,
    employee_id: Optional[str] = None,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Export employee's monthly attendance report with summary stats
    
    - If employee_id is provided, only MANAGER and ADMIN can download
    - If employee_id is not provided, current user gets their own report
    """
    try:
        # Determine which employee to get
        if employee_id:
            # Only MANAGER and ADMIN can download other employees' reports
            if current_user.user_type not in [UserType.MANAGER, UserType.ADMIN]:
                raise HTTPException(status_code=403, detail="Not authorized to download other employees' reports")
            
            # Get the requested employee by employee_id (the ID field, not database id)
            emp_result = await db.execute(
                select(Employee).filter(Employee.employee_id == employee_id)
            )
            employee = emp_result.scalar_one_or_none()
            if not employee:
                raise HTTPException(status_code=404, detail=f"Employee with ID {employee_id} not found")
            
            # For MANAGER, check if employee belongs to their department
            if current_user.user_type == UserType.MANAGER:
                mgr_result = await db.execute(
                    select(Manager).filter(Manager.user_id == current_user.id)
                )
                manager = mgr_result.scalar_one_or_none()
                if manager and employee.department_id != manager.department_id:
                    raise HTTPException(status_code=403, detail="Can only download reports for employees in your department")
        else:
            # Get current user's employee record
            emp_result = await db.execute(
                select(Employee).filter(Employee.user_id == current_user.id)
            )
            employee = emp_result.scalar_one_or_none()
            if not employee:
                raise HTTPException(status_code=404, detail="Employee not found")
        
        # Get attendance for the month
        start_date = date(year, month, 1)
        end_date = date(year, month, monthrange(year, month)[1])
        
        att_result = await db.execute(
            select(Attendance).filter(
                Attendance.employee_id == employee.id,
                Attendance.date >= start_date,
                Attendance.date <= end_date
            ).order_by(Attendance.date)
        )
        attendance_records = att_result.scalars().all()
        
        # Get leave records for the month
        leave_result = await db.execute(
            select(LeaveRequest).filter(
                LeaveRequest.employee_id == employee.id,
                LeaveRequest.start_date <= end_date,
                LeaveRequest.end_date >= start_date,
                LeaveRequest.status == LeaveStatus.APPROVED
            )
        )
        leave_records = leave_result.scalars().all()
        
        # Get schedules for shift information
        sched_result = await db.execute(
            select(Schedule).filter(
                Schedule.employee_id == employee.id,
                Schedule.date >= start_date,
                Schedule.date <= end_date
            )
        )
        schedules = sched_result.scalars().all()
        schedule_map = {}
        for sched in schedules:
            schedule_map[sched.date] = sched
        
        # Calculate leave dates
        leave_dates = set()
        for leave in leave_records:
            current = leave.start_date
            while current <= leave.end_date:
                leave_dates.add(current)
                current += timedelta(days=1)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Monthly Attendance"
        
        # Header styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        summary_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        summary_font = Font(bold=True)
        
        # Title
        ws['A1'] = f"{employee.first_name} {employee.last_name} - Monthly Attendance Report"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:J1')
        
        ws['A2'] = f"{calendar.month_name[month]} {year}"
        ws.merge_cells('A2:J2')
        
        ws['A3'] = f"Employee ID: {employee.employee_id}"
        
        # Headers
        headers = ['Date', 'Day', 'Assigned Shift', 'Check-In', 'Check-Out', 'Hours Worked', 'Break (min)', 'Overtime Hours', 'Status', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data
        row = 6
        total_worked_hours = 0
        total_ot_hours = 0
        working_days_count = 0
        
        for record in attendance_records:
            schedule = schedule_map.get(record.date)
            assigned_shift = '-'
            if schedule and schedule.start_time and schedule.end_time:
                assigned_shift = f"{schedule.start_time} - {schedule.end_time}"
            
            day_name = record.date.strftime('%A')
            
            ws.cell(row=row, column=1).value = record.date.isoformat()
            ws.cell(row=row, column=2).value = day_name
            ws.cell(row=row, column=3).value = assigned_shift
            ws.cell(row=row, column=4).value = record.in_time or '-'
            ws.cell(row=row, column=5).value = record.out_time or '-'
            ws.cell(row=row, column=6).value = f"{record.worked_hours:.2f}" if record.worked_hours else '-'
            ws.cell(row=row, column=7).value = f"{record.break_minutes}" if record.break_minutes else '-'
            ws.cell(row=row, column=8).value = f"{record.overtime_hours:.2f}" if record.overtime_hours else '-'
            ws.cell(row=row, column=9).value = record.status or '-'
            ws.cell(row=row, column=10).value = record.notes or '-'
            
            for col in range(1, 11):
                ws.cell(row=row, column=col).border = border
            
            if record.worked_hours and record.worked_hours > 0:
                total_worked_hours += record.worked_hours
                working_days_count += 1
            
            if record.overtime_hours:
                total_ot_hours += record.overtime_hours
            
            row += 1
        
        # Summary section
        summary_row = row + 2
        
        ws.cell(row=summary_row, column=1).value = "SUMMARY"
        ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
        
        summary_row += 1
        
        # Summary items
        summary_items = [
            ("Working Days Worked", working_days_count),
            ("Leave Days Taken", len(leave_dates)),
            ("Total Hours Worked", f"{total_worked_hours:.2f}"),
            ("Total OT Hours", f"{total_ot_hours:.2f}"),
        ]
        
        for label, value in summary_items:
            ws.cell(row=summary_row, column=1).value = label
            ws.cell(row=summary_row, column=1).font = summary_font
            ws.cell(row=summary_row, column=1).fill = summary_fill
            ws.cell(row=summary_row, column=1).border = border
            
            ws.cell(row=summary_row, column=2).value = value
            ws.cell(row=summary_row, column=2).fill = summary_fill
            ws.cell(row=summary_row, column=2).border = border
            
            summary_row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 20
        
        # Save to bytes
        file_bytes = io.BytesIO()
        wb.save(file_bytes)
        file_bytes.seek(0)
        
        return StreamingResponse(
            iter([file_bytes.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={employee.employee_id}_{employee.first_name}_{year}-{month:02d}_attendance.xlsx"}
        )
    except Exception as e:
        print(f"Employee export error: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    # Save to bytes
    file_bytes = io.BytesIO()
    wb.save(file_bytes)
    file_bytes.seek(0)
    
    return StreamingResponse(
        iter([file_bytes.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={department.name}_attendance_{start_date.isoformat()}_to_{end_date.isoformat()}.xlsx"}
    )


# Leave Requests
@app.post("/leave-requests", response_model=LeaveRequestResponse)
async def create_leave_request(
    leave_data: LeaveRequestCreate,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    # Validate that required fields are provided
    if not leave_data.start_date:
        raise HTTPException(status_code=400, detail="Leave request must have a start date")
    if not leave_data.end_date:
        raise HTTPException(status_code=400, detail="Leave request must have an end date")
    if not leave_data.leave_type:
        raise HTTPException(status_code=400, detail="Leave request must have a leave type")
    
    # Validate date range
    if leave_data.end_date < leave_data.start_date:
        raise HTTPException(status_code=400, detail="End date must be after start date")
    
    # Employee can only request for themselves
    if current_user.user_type == UserType.EMPLOYEE:
        # Find employee record linked to this user
        result = await db.execute(select(Employee).filter(Employee.user_id == current_user.id))
        employee = result.scalar_one_or_none()
        if not employee or leave_data.employee_id != employee.id:
            raise HTTPException(status_code=403, detail="Can only request leave for yourself")
    
    # Get the employee to check paid leave limit
    emp_result = await db.execute(select(Employee).filter(Employee.id == leave_data.employee_id))
    employee = emp_result.scalar_one_or_none()
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # If requesting paid leave, check if it exceeds the annual entitlement
    if leave_data.leave_type == 'paid':
        # Calculate days for this request
        days_requested = (leave_data.end_date - leave_data.start_date).days + 1
        
        # Get already approved paid leave
        # Calculate days as: (end_date - start_date) + 1
        approved_paid_result = await db.execute(
            select(func.sum(
                func.cast(
                    LeaveRequest.end_date - LeaveRequest.start_date + 1,
                    Integer
                )
            )).filter(
                LeaveRequest.employee_id == leave_data.employee_id,
                LeaveRequest.leave_type == 'paid',
                LeaveRequest.status == LeaveStatus.APPROVED
            )
        )
        already_taken = approved_paid_result.scalar() or 0
        
        total_would_be = already_taken + days_requested
        annual_entitlement = employee.paid_leave_per_year
        
        if total_would_be > annual_entitlement:
            remaining = max(0, annual_entitlement - already_taken)
            raise HTTPException(
                status_code=400,
                detail=f"Cannot take {days_requested} days of paid leave. Annual entitlement is {annual_entitlement} days, already taken {already_taken} days. Only {remaining} days remaining."
            )
    
    leave_request = LeaveRequest(**leave_data.dict())
    db.add(leave_request)
    await db.commit()
    await db.refresh(leave_request)
    
    return leave_request


@app.get("/leave-requests", response_model=List[LeaveRequestResponse])
async def list_leave_requests(
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    if current_user.user_type == UserType.EMPLOYEE:
        # Get employee record for current user
        emp_result = await db.execute(
            select(Employee).filter(Employee.user_id == current_user.id)
        )
        employee = emp_result.scalar_one_or_none()
        if not employee:
            return []
        
        result = await db.execute(
            select(LeaveRequest)
            .filter(LeaveRequest.employee_id == employee.id)
            .order_by(LeaveRequest.start_date)
        )
    elif current_user.user_type == UserType.MANAGER:
        # Get all leave requests for employees in manager's department
        manager_dept = await get_manager_department(current_user, db)
        if not manager_dept:
            return []
        
        result = await db.execute(
            select(LeaveRequest)
            .join(Employee)
            .filter(Employee.department_id == manager_dept)
            .order_by(LeaveRequest.start_date)
        )
    else:  # Admin
        result = await db.execute(
            select(LeaveRequest)
            .order_by(LeaveRequest.start_date)
        )
    
    return result.scalars().all()


@app.get("/leave-statistics")
async def get_leave_statistics(
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Get leave statistics for current employee (or all if manager/admin)"""
    if current_user.user_type == UserType.EMPLOYEE:
        # Get employee record for current user
        emp_result = await db.execute(
            select(Employee).filter(Employee.user_id == current_user.id)
        )
        employee = emp_result.scalar_one_or_none()
        if not employee:
            raise HTTPException(status_code=404, detail="Employee not found")
        
        # Get all leave requests for this employee
        result = await db.execute(
            select(LeaveRequest)
            .filter(LeaveRequest.employee_id == employee.id, LeaveRequest.status == LeaveStatus.APPROVED)
        )
        approved_leaves = result.scalars().all()
        
        # Calculate taken paid leaves
        from datetime import date
        taken_paid = 0
        for leave in approved_leaves:
            if leave.leave_type == 'paid':
                days = (leave.end_date - leave.start_date).days + 1
                taken_paid += days
        
        total_paid_leave = employee.paid_leave_per_year  # Use employee's paid leave setting
        available_paid = max(0, total_paid_leave - taken_paid)
        
        return {
            "total_paid_leave": total_paid_leave,
            "taken_paid_leave": taken_paid,
            "available_paid_leave": available_paid,
            "employee_name": f"{employee.first_name} {employee.last_name}"
        }
    else:
        raise HTTPException(status_code=403, detail="Only employees can access their statistics")


@app.get("/leave-statistics/employee/{employee_id}")
async def get_employee_leave_statistics(
    employee_id: str,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    """Get leave statistics for a specific employee (manager only) with monthly breakdown"""
    from datetime import date, datetime
    from collections import defaultdict
    
    # Get the manager record for current user
    manager_result = await db.execute(select(Manager).filter(Manager.user_id == current_user.id))
    manager = manager_result.scalar_one_or_none()
    
    if not manager:
        raise HTTPException(status_code=403, detail="User is not a manager")
    
    # Get employee record by employee_id (string field)
    emp_result = await db.execute(
        select(Employee).filter(Employee.employee_id == employee_id, Employee.department_id == manager.department_id)
    )
    employee = emp_result.scalar_one_or_none()
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found in your department")
    
    # Get all approved leave requests for this employee using the integer id
    result = await db.execute(
        select(LeaveRequest)
        .filter(LeaveRequest.employee_id == employee.id, LeaveRequest.status == LeaveStatus.APPROVED)
        .order_by(LeaveRequest.start_date)
    )
    approved_leaves = result.scalars().all()
    
    # Calculate leave statistics and monthly breakdown
    taken_paid = 0
    taken_unpaid = 0
    monthly_breakdown = defaultdict(lambda: {'paid': 0, 'unpaid': 0, 'total': 0})
    
    for leave in approved_leaves:
        days = (leave.end_date - leave.start_date).days + 1
        month_key = leave.start_date.strftime('%Y-%m')  # Format: "2025-01"
        month_name = leave.start_date.strftime('%B %Y')  # Format: "January 2025"
        
        if leave.leave_type == 'paid':
            taken_paid += days
            monthly_breakdown[month_key]['paid'] += days
        else:
            taken_unpaid += days
            monthly_breakdown[month_key]['unpaid'] += days
        
        monthly_breakdown[month_key]['total'] += days
        monthly_breakdown[month_key]['month_name'] = month_name
    
    total_paid_leave = employee.paid_leave_per_year  # Use employee's paid leave setting
    available_paid = max(0, total_paid_leave - taken_paid)
    
    # Convert monthly breakdown to list and sort by month
    monthly_list = []
    for month_key in sorted(monthly_breakdown.keys()):
        monthly_list.append({
            'month': monthly_breakdown[month_key]['month_name'],
            'paid': monthly_breakdown[month_key]['paid'],
            'unpaid': monthly_breakdown[month_key]['unpaid'],
            'total': monthly_breakdown[month_key]['total']
        })
    
    return {
        "employee_id": employee.employee_id,
        "employee_name": f"{employee.first_name} {employee.last_name}",
        "total_paid_leave": total_paid_leave,
        "taken_paid_leave": taken_paid,
        "taken_unpaid_leave": taken_unpaid,
        "available_paid_leave": available_paid,
        "total_leaves_taken": taken_paid + taken_unpaid,
        "monthly_breakdown": monthly_list
    }


@app.post("/manager/approve-leave/{leave_id}")
async def approve_leave(
    leave_id: int,
    approval_data: LeaveApproval,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    from app.helpers import calculate_leave_days, deduct_leave_balance, create_notification
    
    result = await db.execute(select(LeaveRequest).filter(LeaveRequest.id == leave_id))
    leave_request = result.scalar_one_or_none()

    if not leave_request:
        raise HTTPException(status_code=404, detail="Leave request not found")

    # Get the manager record for current user
    manager_result = await db.execute(select(Manager).filter(Manager.user_id == current_user.id))
    manager = manager_result.scalar_one_or_none()

    if not manager:
        raise HTTPException(status_code=403, detail="User is not a manager")

    # ========== CALCULATE AND DEDUCT LEAVE BALANCE ==========
    days_to_deduct = await calculate_leave_days(
        leave_request.start_date,
        leave_request.end_date,
        leave_request.leave_type
    )
    
    success, message = await deduct_leave_balance(
        db,
        leave_request.employee_id,
        days_to_deduct,
        leave_request.leave_type
    )
    
    if not success:
        # Balance deduction failed, reject the leave
        leave_request.status = LeaveStatus.REJECTED
        leave_request.manager_id = manager.id
        leave_request.reviewed_at = datetime.utcnow()
        leave_request.review_notes = f"Auto-rejected: {message}"
        await db.commit()
        raise HTTPException(status_code=400, detail=f"Cannot approve: {message}")
    
    # ========== APPROVE THE LEAVE ==========
    leave_request.status = LeaveStatus.APPROVED
    leave_request.manager_id = manager.id
    leave_request.reviewed_at = datetime.utcnow()
    leave_request.review_notes = approval_data.review_notes
    
    # ========== GET EMPLOYEE INFO FOR NOTIFICATION ==========
    emp_result = await db.execute(select(Employee).where(Employee.id == leave_request.employee_id))
    employee = emp_result.scalar_one_or_none()
    
    # ========== CREATE NOTIFICATION FOR EMPLOYEE ==========
    if employee and employee.user_id:
        await create_notification(
            db,
            employee.user_id,
            "Leave Request Approved",
            f"Your leave request from {leave_request.start_date} to {leave_request.end_date} has been approved. ({days_to_deduct} days)",
            notification_type="leave_approved",
            related_id=leave_request.id
        )
    
    await db.commit()

    return {
        "message": "Leave approved successfully",
        "days_deducted": days_to_deduct,
        "leave_balance_update": message
    }


@app.post("/manager/reject-leave/{leave_id}")
async def reject_leave(
    leave_id: int,
    approval_data: LeaveApproval,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    from app.helpers import create_notification
    
    result = await db.execute(select(LeaveRequest).filter(LeaveRequest.id == leave_id))
    leave_request = result.scalar_one_or_none()

    if not leave_request:
        raise HTTPException(status_code=404, detail="Leave request not found")

    # Get the manager record for current user
    manager_result = await db.execute(select(Manager).filter(Manager.user_id == current_user.id))
    manager = manager_result.scalar_one_or_none()

    if not manager:
        raise HTTPException(status_code=403, detail="User is not a manager")

    leave_request.status = LeaveStatus.REJECTED
    leave_request.manager_id = manager.id
    leave_request.reviewed_at = datetime.utcnow()
    leave_request.review_notes = approval_data.review_notes
    
    # ========== GET EMPLOYEE INFO FOR NOTIFICATION ==========
    emp_result = await db.execute(select(Employee).where(Employee.id == leave_request.employee_id))
    employee = emp_result.scalar_one_or_none()
    
    # ========== CREATE NOTIFICATION FOR EMPLOYEE ==========
    if employee and employee.user_id:
        await create_notification(
            db,
            employee.user_id,
            "Leave Request Rejected",
            f"Your leave request from {leave_request.start_date} to {leave_request.end_date} has been rejected.\nReason: {approval_data.review_notes}",
            notification_type="leave_rejected",
            related_id=leave_request.id
        )

    await db.commit()

    return {"message": "Leave rejected"}


# Messages
@app.post("/messages", response_model=MessageResponse)
async def send_message(
    message_data: MessageCreate,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    message = Message(
        sender_id=current_user.id,
        recipient_id=message_data.recipient_id,
        department_id=message_data.department_id,
        subject=message_data.subject,
        message=message_data.message
    )
    
    db.add(message)
    await db.commit()
    
    # Refresh and eagerly load relationships
    await db.refresh(message, ['sender', 'recipient'])
    
    return message


@app.get("/messages", response_model=List[MessageResponse])
async def get_messages(
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    # Get messages: either sent by user, sent to user, or sent to their department
    user_department_id = await get_user_department(current_user, db)
    department_filters = []
    if user_department_id is not None:
        department_filters.append(
            and_(
                Message.department_id == user_department_id,
                Message.is_deleted_by_recipient == False
            )
        )
    
    result = await db.execute(
        select(Message).options(
            selectinload(Message.sender),
            selectinload(Message.recipient)
        ).filter(
            or_(
                and_(
                    Message.sender_id == current_user.id,  # Messages sent by the user
                    Message.is_deleted_by_sender == False
                ),
                and_(
                    Message.recipient_id == current_user.id,
                    Message.is_deleted_by_recipient == False
                ),
                *department_filters
            )
        ).order_by(Message.created_at.desc())
    )
    
    return result.scalars().all()


@app.delete("/messages/{message_id}")
async def delete_message(
    message_id: int,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(select(Message).filter(Message.id == message_id))
    message = result.scalar_one_or_none()
    
    if not message:
        raise HTTPException(status_code=404, detail="Message not found")
    
    # Soft delete based on user role
    if message.sender_id == current_user.id:
        message.is_deleted_by_sender = True
    elif message.recipient_id == current_user.id:
        message.is_deleted_by_recipient = True
    else:
        raise HTTPException(status_code=403, detail="Cannot delete this message")
    
    await db.commit()

    return {"message": "Message deleted"}


@app.put("/messages/{message_id}/read")
async def mark_message_as_read(
    message_id: int,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(select(Message).filter(Message.id == message_id))
    message = result.scalar_one_or_none()
    
    if not message:
        raise HTTPException(status_code=404, detail="Message not found")
    
    # Only the recipient or department can mark as read
    user_department_id = await get_user_department(current_user, db)
    if message.recipient_id != current_user.id and message.department_id != user_department_id:
        raise HTTPException(status_code=403, detail="Cannot mark this message as read")
    
    message.is_read = True
    message.read_at = datetime.utcnow()
    await db.commit()
    
    return {"message": "Message marked as read"}


# Notifications
@app.get("/notifications", response_model=List[NotificationResponse])
async def get_notifications(
    unread_only: bool = False,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    query = select(Notification).filter(Notification.user_id == current_user.id)

    if unread_only:
        query = query.filter(Notification.is_read == False)

    result = await db.execute(query.order_by(Notification.created_at.desc()))
    return result.scalars().all()


@app.post("/notifications/{notification_id}/mark-read")
async def mark_notification_read(
    notification_id: int,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(Notification).filter(
            Notification.id == notification_id,
            Notification.user_id == current_user.id
        )
    )
    notification = result.scalar_one_or_none()

    if not notification:
        raise HTTPException(status_code=404, detail="Notification not found")

    notification.is_read = True
    await db.commit()

    return {"message": "Notification marked as read"}


@app.post("/notifications/mark-all-read")
async def mark_all_notifications_read(
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(Notification).filter(
            Notification.user_id == current_user.id,
            Notification.is_read == False
        )
    )
    notifications = result.scalars().all()

    for notification in notifications:
        notification.is_read = True

    await db.commit()

    return {"message": f"{len(notifications)} notifications marked as read"}


@app.delete("/notifications/{notification_id}")
async def delete_notification(
    notification_id: int,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(Notification).filter(
            Notification.id == notification_id,
            Notification.user_id == current_user.id
        )
    )
    notification = result.scalar_one_or_none()

    if not notification:
        raise HTTPException(status_code=404, detail="Notification not found")

    await db.delete(notification)
    await db.commit()

    return {"message": "Notification deleted"}


# Schedules
@app.get("/schedules", response_model=List[ScheduleResponse])
async def get_schedules(
    start_date: date = None,
    end_date: date = None,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    query = select(Schedule).options(
        selectinload(Schedule.employee).selectinload(Employee.user),
        selectinload(Schedule.role)
    )

    if current_user.user_type == UserType.EMPLOYEE:
        # Get employee by user_id
        emp_result = await db.execute(
            select(Employee).filter(Employee.user_id == current_user.id)
        )
        employee = emp_result.scalar_one_or_none()
        if employee:
            query = query.filter(Schedule.employee_id == employee.id)
        else:
            return []
    elif current_user.user_type == UserType.MANAGER:
        manager_dept = await get_manager_department(current_user, db)
        if manager_dept:
            query = query.filter(Schedule.department_id == manager_dept)
        else:
            return []

    if start_date:
        query = query.filter(Schedule.date >= start_date)
    if end_date:
        query = query.filter(Schedule.date <= end_date)

    result = await db.execute(query.order_by(Schedule.date))
    return result.scalars().all()


@app.post("/schedules", response_model=ScheduleResponse)
async def create_schedule(
    schedule_data: ScheduleCreate,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    # Get employee to verify department
    result = await db.execute(select(Employee).filter(Employee.id == schedule_data.employee_id))
    employee = result.scalar_one_or_none()

    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")

    if current_user.user_type == UserType.MANAGER:
        manager_dept = await get_manager_department(current_user, db)
        if not manager_dept or employee.department_id != manager_dept:
            raise HTTPException(status_code=403, detail="Can only schedule employees in your department")

    # Get the shift/role to calculate hours
    try:
        start_h, start_m = map(int, schedule_data.start_time.split(':'))
        end_h, end_m = map(int, schedule_data.end_time.split(':'))
        start_decimal = start_h + start_m / 60
        end_decimal = end_h + end_m / 60
        shift_hours = end_decimal - start_decimal if end_decimal > start_decimal else 24 - start_decimal + end_decimal
    except:
        shift_hours = 0
    
    # CONSTRAINT 1: Check if assigning more than 5 shifts per week
    week_start = schedule_data.date - timedelta(days=schedule_data.date.weekday())
    week_end = week_start + timedelta(days=6)
    
    week_count_result = await db.execute(
        select(func.count(Schedule.id)).filter(
            Schedule.employee_id == schedule_data.employee_id,
            Schedule.date >= week_start,
            Schedule.date <= week_end
        )
    )
    existing_shifts_count = week_count_result.scalar() or 0
    
    if existing_shifts_count >= 5:
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot assign more than 5 shifts per week. Employee already has {existing_shifts_count} shifts this week (max: 5)"
        )
    
    # ===== NEW CONSTRAINT: Check 5 consecutive shifts limit =====
    # Check if this new shift would create more than 5 consecutive shifts
    week_schedules_check = await db.execute(
        select(Schedule).filter(
            Schedule.employee_id == schedule_data.employee_id,
            Schedule.date >= week_start,
            Schedule.date <= week_end
        ).order_by(Schedule.date)
    )
    existing_week_schedules = week_schedules_check.scalars().all()
    
    # Collect all dates including the new one
    all_dates = sorted([s.date for s in existing_week_schedules] + [schedule_data.date])
    all_dates = list(dict.fromkeys(all_dates))  # Remove duplicates while preserving order
    
    # Find the longest consecutive sequence
    max_consecutive = 1
    current_consecutive = 1
    for i in range(1, len(all_dates)):
        if (all_dates[i] - all_dates[i-1]).days == 1:
            current_consecutive += 1
            max_consecutive = max(max_consecutive, current_consecutive)
        else:
            current_consecutive = 1
    
    if max_consecutive > 5:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot create {max_consecutive} consecutive shifts. Maximum allowed is 5 consecutive shifts."
        )
    
    # Check if this creates overtime (exceeds 8hrs/day)
    overtime_required = False
    overtime_hours = 0
    daily_overtime = False
    weekly_overtime = False
    
    if shift_hours > 8:
        overtime_required = True
        daily_overtime = True
        overtime_hours = shift_hours - 8
    
    # Check existing schedules for the day to see if daily limit exceeded
    existing_result = await db.execute(
        select(Schedule).filter(
            Schedule.employee_id == schedule_data.employee_id,
            Schedule.date == schedule_data.date
        )
    )
    existing_schedules = existing_result.scalars().all()
    
    existing_hours = 0
    for sched in existing_schedules:
        try:
            start_h, start_m = map(int, sched.start_time.split(':'))
            end_h, end_m = map(int, sched.end_time.split(':'))
            start_decimal = start_h + start_m / 60
            end_decimal = end_h + end_m / 60
            hours = end_decimal - start_decimal if end_decimal > start_decimal else 24 - start_decimal + end_decimal
            existing_hours += hours
        except:
            pass
    
    # CONSTRAINT 2: Total hours for the day must be 9 hrs (8 hrs work + 1 hr break)
    total_daily_hours = existing_hours + shift_hours
    
    if total_daily_hours > 8:
        overtime_required = True
        daily_overtime = True
        if not daily_overtime or shift_hours <= 8:
            overtime_hours = total_daily_hours - 8
    
    # CONSTRAINT 3: Check weekly hours - max 40 hours per week
    week_result = await db.execute(
        select(Schedule).filter(
            Schedule.employee_id == schedule_data.employee_id,
            Schedule.date >= week_start,
            Schedule.date <= week_end
        )
    )
    week_schedules = week_result.scalars().all()
    
    existing_weekly_hours = 0
    for sched in week_schedules:
        try:
            start_h, start_m = map(int, sched.start_time.split(':'))
            end_h, end_m = map(int, sched.end_time.split(':'))
            start_decimal = start_h + start_m / 60
            end_decimal = end_h + end_m / 60
            hours = end_decimal - start_decimal if end_decimal > start_decimal else 24 - start_decimal + end_decimal
            existing_weekly_hours += hours
        except:
            pass
    
    total_weekly_hours = existing_weekly_hours + shift_hours
    weekly_overtime_hours = 0
    
    if total_weekly_hours > 40:
        overtime_required = True
        weekly_overtime = True
        weekly_overtime_hours = total_weekly_hours - 40
    
    # If overtime required, return info about it (frontend will show popup)
    if overtime_required:
        # Check if employee has enough OT available
        year = schedule_data.date.year
        month = schedule_data.date.month
        
        tracking_result = await db.execute(
            select(OvertimeTracking).filter(
                OvertimeTracking.employee_id == schedule_data.employee_id,
                OvertimeTracking.year == year,
                OvertimeTracking.month == month
            )
        )
        tracking = tracking_result.scalar_one_or_none()
        
        if not tracking:
            # Create default tracking
            tracking = OvertimeTracking(
                employee_id=schedule_data.employee_id,
                year=year,
                month=month,
                allocated_hours=8,
                used_hours=0.0,
                remaining_hours=8
            )
            db.add(tracking)
            await db.commit()
            await db.refresh(tracking)
        
        total_overtime_hours = max(overtime_hours, weekly_overtime_hours)
        has_sufficient_ot = tracking.remaining_hours >= total_overtime_hours
        
        messages = []
        if daily_overtime:
            messages.append(f"Daily overtime: {overtime_hours:.1f}h")
        if weekly_overtime:
            messages.append(f"Weekly overtime: {weekly_overtime_hours:.1f}h")
        
        return {
            "id": None,
            "status": "requires_overtime_approval",
            "message": f"Overtime required: {', '.join(messages)}",
            "employee_id": schedule_data.employee_id,
            "employee_name": f"{employee.first_name} {employee.last_name}",
            "date": schedule_data.date.isoformat(),
            "start_time": schedule_data.start_time,
            "end_time": schedule_data.end_time,
            "shift_hours": shift_hours,
            "overtime_hours": total_overtime_hours,
            "total_daily_hours": total_daily_hours if total_daily_hours > 8 else shift_hours,
            "total_weekly_hours": total_weekly_hours,
            "allocated_ot_hours": tracking.allocated_hours,
            "used_ot_hours": tracking.used_hours,
            "remaining_ot_hours": tracking.remaining_hours,
            "has_sufficient_ot": has_sufficient_ot,
            "daily_overtime": daily_overtime,
            "weekly_overtime": weekly_overtime
        }
    
    # Create schedule normally if no overtime required
    schedule = Schedule(
        department_id=employee.department_id,
        employee_id=schedule_data.employee_id,
        role_id=schedule_data.role_id,
        shift_id=schedule_data.shift_id,  # Optional - can be None for custom schedules
        date=schedule_data.date,
        start_time=schedule_data.start_time,
        end_time=schedule_data.end_time,
        notes=schedule_data.notes,
        status='scheduled'
    )

    db.add(schedule)
    await db.commit()

    # Refresh with eager loading
    result = await db.execute(
        select(Schedule)
        .filter(Schedule.id == schedule.id)
        .options(selectinload(Schedule.role))
    )
    return result.scalar_one()


@app.put("/schedules/{schedule_id}", response_model=ScheduleResponse)
async def update_schedule(
    schedule_id: int,
    schedule_data: ScheduleUpdate,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(select(Schedule).filter(Schedule.id == schedule_id))
    schedule = result.scalar_one_or_none()

    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")

    if current_user.user_type == UserType.MANAGER:
        manager_dept = await get_manager_department(current_user, db)
        if not manager_dept or schedule.department_id != manager_dept:
            raise HTTPException(status_code=403, detail="Can only edit schedules in your department")

    # ===== NEW: Validate consecutive shifts when changing date =====
    if schedule_data.date and schedule_data.date != schedule.date:
        new_date = schedule_data.date if isinstance(schedule_data.date, date) else datetime.strptime(schedule_data.date, '%Y-%m-%d').date()
        
        week_start = new_date - timedelta(days=new_date.weekday())
        week_end = week_start + timedelta(days=6)
        
        week_schedules_check = await db.execute(
            select(Schedule).filter(
                Schedule.employee_id == schedule.employee_id,
                Schedule.date >= week_start,
                Schedule.date <= week_end,
                Schedule.id != schedule_id  # Exclude current schedule
            ).order_by(Schedule.date)
        )
        existing_week_schedules = week_schedules_check.scalars().all()
        
        # Collect all dates including the new one
        all_dates = sorted([s.date for s in existing_week_schedules] + [new_date])
        all_dates = list(dict.fromkeys(all_dates))  # Remove duplicates while preserving order
        
        # Find the longest consecutive sequence
        max_consecutive = 1
        current_consecutive = 1
        for i in range(1, len(all_dates)):
            if (all_dates[i] - all_dates[i-1]).days == 1:
                current_consecutive += 1
                max_consecutive = max(max_consecutive, current_consecutive)
            else:
                current_consecutive = 1
        
        if max_consecutive > 5:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot create {max_consecutive} consecutive shifts. Maximum allowed is 5 consecutive shifts."
            )

    for key, value in schedule_data.dict(exclude_unset=True).items():
        # Convert string date to date object if needed
        if key == 'date' and isinstance(value, str):
            value = datetime.strptime(value, '%Y-%m-%d').date()
        setattr(schedule, key, value)

    schedule.updated_at = datetime.utcnow()
    await db.commit()

    # Re-fetch with eager loading
    result = await db.execute(
        select(Schedule)
        .filter(Schedule.id == schedule_id)
        .options(selectinload(Schedule.role))
    )
    return result.scalar_one()


@app.delete("/schedules/{schedule_id}")
async def delete_schedule(
    schedule_id: int,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(select(Schedule).filter(Schedule.id == schedule_id))
    schedule = result.scalar_one_or_none()

    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")

    if current_user.user_type == UserType.MANAGER:
        manager_dept = await get_manager_department(current_user, db)
        if not manager_dept or schedule.department_id != manager_dept:
            raise HTTPException(status_code=403, detail="Can only delete schedules in your department")

    await db.delete(schedule)
    await db.commit()

    return {"message": "Schedule deleted successfully"}


@app.post("/schedules/generate")
async def generate_schedules(
    start_date: date,
    end_date: date,
    regenerate: bool = False,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    """
    Generate fair schedules with equal shift distribution.

    Algorithm:
    1. Get all roles and shifts for the department in the date range
    2. Get all employees in the department
    3. Calculate each employee's capacity (shifts per week)
    4. Fairly assign shifts equally across different shift types
    5. Respect min_emp and max_emp constraints for each shift
    """
    try:
        print(f"[DEBUG] Schedule generation started for dates {start_date} to {end_date}", flush=True)

        # Get manager's department
        department_id = await get_manager_department(current_user, db)
        if not department_id:
            raise HTTPException(status_code=400, detail="Manager department not found")

        print(f"[DEBUG] Department ID: {department_id}", flush=True)

        # ===== NEW: Check if schedules already exist in this date range =====
        existing_schedules_result = await db.execute(
            select(Schedule)
            .filter(
                Schedule.department_id == department_id,
                Schedule.date >= start_date,
                Schedule.date <= end_date
            )
        )
        existing_schedules = existing_schedules_result.scalars().all()
        
        if existing_schedules and not regenerate:
            # Return message asking if user wants to regenerate
            return {
                "success": False,
                "schedules_created": 0,
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "requires_confirmation": True,
                "existing_count": len(existing_schedules),
                "feedback": [
                    f"⚠️  Found {len(existing_schedules)} existing schedules for this date range.",
                    "Do you want to regenerate and replace them?"
                ],
                "schedules": []
            }
        
        # If regenerate is True, delete existing schedules first
        if existing_schedules and regenerate:
            print(f"[DEBUG] Regenerating - deleting {len(existing_schedules)} existing schedules", flush=True)
            
            # First, nullify schedule_id in check_ins table for affected date range
            # to avoid foreign key constraint violations
            affected_schedules_result = await db.execute(
                select(Schedule.id)
                .filter(
                    Schedule.department_id == department_id,
                    Schedule.date >= start_date,
                    Schedule.date <= end_date
                )
            )
            affected_schedule_ids = affected_schedules_result.scalars().all()
            
            if affected_schedule_ids:
                await db.execute(
                    update(CheckInOut)
                    .where(CheckInOut.schedule_id.in_(affected_schedule_ids))
                    .values(schedule_id=None)
                )
            
            # Now delete the schedules
            await db.execute(
                delete(Schedule)
                .filter(
                    Schedule.department_id == department_id,
                    Schedule.date >= start_date,
                    Schedule.date <= end_date
                )
            )
            await db.commit()
            feedback = [f"Cleared {len(existing_schedules)} existing schedules. Generating new schedule..."]
        else:
            feedback = []

        # Get all roles in this department
        roles_result = await db.execute(
            select(Role)
            .filter(Role.department_id == department_id, Role.is_active == True)
        )
        roles = roles_result.scalars().all()
        print(f"[DEBUG] Found {len(roles)} roles", flush=True)

        if not roles:
            return {
                "success": True,
                "schedules_created": 0,
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "feedback": ["❌ No active roles found in your department. Create roles first."],
                "schedules": []
            }

        # Get all shifts for these roles
        role_ids = [r.id for r in roles]
        shifts_result = await db.execute(
            select(Shift)
            .filter(Shift.role_id.in_(role_ids), Shift.is_active == True)
        )
        shifts = shifts_result.scalars().all()
        print(f"[DEBUG] Found {len(shifts)} shifts", flush=True)

        # Log shift details and ensure all shifts have schedule_config
        print(f"[DEBUG] Processing {len(shifts)} shifts for schedule_config validation", flush=True)
        for shift in shifts:
            # For backward compatibility:
            # - If shift has NO schedule_config or empty, assume ALL days are enabled
            # - If shift has schedule_config, use the configured days
            if not shift.schedule_config or not isinstance(shift.schedule_config, dict) or len(shift.schedule_config) == 0:
                print(f"[DEBUG] Shift {shift.id} ({shift.name}) has empty/invalid schedule_config, enabling all days for backward compatibility", flush=True)
                # Old shift without schedule_config - enable all days for backward compatibility
                shift.schedule_config = {
                    'Monday': {'enabled': True},
                    'Tuesday': {'enabled': True},
                    'Wednesday': {'enabled': True},
                    'Thursday': {'enabled': True},
                    'Friday': {'enabled': True},
                    'Saturday': {'enabled': True},
                    'Sunday': {'enabled': True}
                }
            else:
                # Ensure all days have proper structure
                if isinstance(shift.schedule_config, dict):
                    for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']:
                        if day not in shift.schedule_config or not isinstance(shift.schedule_config[day], dict):
                            # Missing day or malformed - fix it
                            shift.schedule_config[day] = {'enabled': False}
                        elif 'enabled' not in shift.schedule_config[day]:
                            # Missing 'enabled' key - add it
                            shift.schedule_config[day]['enabled'] = False
            
            enabled_days = [day for day, cfg in shift.schedule_config.items() if isinstance(cfg, dict) and cfg.get('enabled', False)]
            print(f"[DEBUG] Final Shift: {shift.id} - {shift.name}, enabled_days={enabled_days}", flush=True)

        if not shifts:
            return {
                "success": True,
                "schedules_created": 0,
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "feedback": ["❌ No active shifts found in your roles. Create shifts first."],
                "schedules": []
            }

        # Get all employees in the department
        employees_result = await db.execute(
            select(Employee)
            .filter(Employee.department_id == department_id, Employee.is_active == True)
        )
        employees = employees_result.scalars().all()
        print(f"[DEBUG] Found {len(employees)} employees", flush=True)

        # Log employee details
        for emp in employees:
            print(f"[DEBUG] Employee: {emp.id} - {emp.first_name}, active={emp.is_active}, weekly_hours={emp.weekly_hours}, daily_max={emp.daily_max_hours}, shifts_per_week={emp.shifts_per_week}", flush=True)

        if not employees:
            return {
                "success": True,
                "schedules_created": 0,
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "feedback": ["❌ No active employees found in your department. Create employees first."],
                "schedules": []
            }

        # Generate date range (one schedule per shift per day)
        current_date = start_date
        schedules_created = 0
        feedback = []
        overtime_warnings = []  # Track shifts requiring overtime approval

        # Group shifts by role for fair distribution
        shifts_by_role = defaultdict(list)
        for shift in shifts:
            shifts_by_role[shift.role_id].append(shift)

        # Calculate which employees are eligible for each shift (based on role)
        # Strategy: Assign employees to shifts day-by-day
        # Each employee gets ONE shift per day maximum (no double shifts)
        # If shifts are Mon-Friday, all eligible employees get all 5 days
        eligible_for_shift = {}  # {shift_id: [emp1, emp2, emp3...]} - only eligible employees per shift

        print(f"[DEBUG] Building eligibility matrix for {len(shifts)} shifts and {len(employees)} employees", flush=True)
        for shift in shifts:
            eligible_for_shift[shift.id] = []
            
            for emp in employees:
                # Employee is eligible if:
                # 1. They have no specific role assignment (flexible=True), OR
                # 2. Their role matches the shift's role
                is_eligible = (emp.role_id is None) or (emp.role_id == shift.role_id)
                
                if is_eligible:
                    eligible_for_shift[shift.id].append(emp)
                    print(f"[DEBUG] Shift {shift.id} ({shift.name}): {emp.id} ({emp.first_name}) is ELIGIBLE", flush=True)
                else:
                    print(f"[DEBUG] Shift {shift.id} ({shift.name}): {emp.id} ({emp.first_name}) is NOT eligible (role mismatch: emp.role={emp.role_id} vs shift.role={shift.role_id})", flush=True)

        # Create schedules
        current_date = start_date
        while current_date <= end_date:
            day_name = current_date.strftime('%A')  # e.g., 'Monday', 'Sunday'

            for shift in shifts:
                # Check if shift operates on this day
                role = next((r for r in roles if r.id == shift.role_id), None)
                
                # Determine if this shift should run on this day
                should_skip = False
                
                if shift.schedule_config and isinstance(shift.schedule_config, dict):
                    # Shift has a schedule_config with day configuration
                    day_config = shift.schedule_config.get(day_name, {})
                    is_day_enabled = day_config.get('enabled', False) if isinstance(day_config, dict) else False
                    
                    if not is_day_enabled:
                        should_skip = True
                        print(f"[DEBUG] ✗ Shift {shift.id} ({shift.name}) - Day {day_name} is disabled, skipping", flush=True)
                    else:
                        print(f"[DEBUG] ✓ Shift {shift.id} ({shift.name}) - Day {day_name} is ENABLED, processing", flush=True)
                else:
                    # No schedule_config or invalid format - skip to prevent unintended assignments
                    should_skip = True
                    print(f"[DEBUG] ✗ Shift {shift.id} ({shift.name}) - No valid schedule_config, skipping {day_name}", flush=True)

                if should_skip:
                    continue

                # Assign employees to this shift on this day
                # Only consider employees who are eligible for this shift
                assigned_count = 0
                
                for emp in eligible_for_shift[shift.id]:
                    # Check leave - if employee is on approved leave, mark them as leave (not shift)
                    leave_result = await db.execute(
                        select(LeaveRequest)
                        .filter(
                            LeaveRequest.employee_id == emp.id,
                            LeaveRequest.start_date <= current_date,
                            LeaveRequest.end_date >= current_date,
                            LeaveRequest.status == LeaveStatus.APPROVED
                        )
                    )
                    leave_request = leave_result.scalars().first()
                    
                    if leave_request:
                        # Employee is on approved leave - create a LEAVE schedule entry
                        # (This counts as one of their working days, but marked as leave)
                        existing_today = await db.execute(
                            select(Schedule)
                            .filter(
                                Schedule.employee_id == emp.id,
                                Schedule.date == current_date
                            )
                        )
                        if not existing_today.scalars().first():
                            # No schedule exists for this day yet - create a leave entry
                            print(f"[DEBUG] ✓ {emp.first_name} is on approved leave on {current_date}, creating LEAVE schedule", flush=True)
                            leave_schedule = Schedule(
                                department_id=department_id,
                                employee_id=emp.id,
                                role_id=shift.role_id,
                                shift_id=shift.id,
                                date=current_date,
                                start_time=None,
                                end_time=None,
                                status="leave"  # Mark as leave, not shift
                            )
                            db.add(leave_schedule)
                            schedules_created += 1
                        else:
                            print(f"[DEBUG] ✗ {emp.first_name} already has a schedule entry on {current_date}, skipping leave creation", flush=True)
                        continue  # Don't assign shift for leave day
                    
                    # CRITICAL: Check if employee already has a shift on this day (NO DOUBLE SHIFTS)
                    existing_today = await db.execute(
                        select(Schedule)
                        .filter(
                            Schedule.employee_id == emp.id,
                            Schedule.date == current_date
                        )
                    )
                    if existing_today.scalars().first():
                        print(f"[DEBUG] ✗ {emp.first_name} already has a shift on {current_date}, skipping (NO DOUBLE SHIFTS)", flush=True)
                        continue  # Skip if employee already has a shift today
                    
                    print(f"[DEBUG] Checking {emp.first_name} ({emp.id}) for shift {shift.id} ({shift.name}) on {current_date}", flush=True)
                    
                    # Check 5 consecutive shifts limit
                    week_start = current_date - timedelta(days=current_date.weekday())
                    week_end = week_start + timedelta(days=6)
                    
                    consecutive_check = await db.execute(
                        select(Schedule)
                        .filter(
                            Schedule.employee_id == emp.id,
                            Schedule.date >= week_start,
                            Schedule.date <= week_end
                        )
                        .order_by(Schedule.date)
                    )
                    week_schedules = consecutive_check.scalars().all()
                    
                    # Check consecutive shifts INCLUDING the new one
                    week_dates = [s.date for s in week_schedules]
                    if current_date not in week_dates:
                        week_dates.append(current_date)
                    
                    week_dates.sort()
                    max_consecutive = 1
                    current_consecutive = 1
                    
                    for i in range(1, len(week_dates)):
                        if (week_dates[i] - week_dates[i-1]).days == 1:
                            current_consecutive += 1
                            max_consecutive = max(max_consecutive, current_consecutive)
                        else:
                            current_consecutive = 1
                    
                    if max_consecutive > 5:
                        print(f"[DEBUG] ✗ {emp.first_name} would have {max_consecutive} consecutive shifts, skipping (MAX 5 consecutive)", flush=True)
                        continue  # Skip if would exceed 5 consecutive shifts

                    # Fetch existing schedules for the week (with eager loading of role)
                    existing_schedules_result = await db.execute(
                        select(Schedule)
                        .filter(
                            Schedule.employee_id == emp.id,
                            Schedule.date >= week_start,
                            Schedule.date <= week_end
                        )
                        .options(selectinload(Schedule.role))
                    )
                    existing_schedules = existing_schedules_result.scalars().all()

                    # Calculate existing hours in Python, subtracting break time
                    existing_hours = 0
                    existing_hours_today = 0
                    for sched in existing_schedules:
                        if sched.start_time and sched.end_time:
                            try:
                                start = datetime.strptime(sched.start_time, '%H:%M')
                                end = datetime.strptime(sched.end_time, '%H:%M')
                                total_hours = (end - start).total_seconds() / 3600

                                # Subtract break time from role
                                break_hours = (sched.role.break_minutes or 0) / 60 if sched.role else 0
                                work_hours = total_hours - break_hours

                                existing_hours += work_hours
                                # Check hours for current day
                                if sched.date == current_date:
                                    existing_hours_today += work_hours
                            except (ValueError, TypeError):
                                pass

                    # Calculate shift hours (total time) and work hours (minus breaks)
                    shift_start = datetime.strptime(shift.start_time, '%H:%M')
                    shift_end = datetime.strptime(shift.end_time, '%H:%M')
                    total_shift_hours = (shift_end - shift_start).total_seconds() / 3600

                    # Subtract break time from role
                    break_hours = (role.break_minutes or 0) / 60
                    work_hours = total_shift_hours - break_hours

                    # Check both weekly and daily limits using work hours (excluding breaks)
                    daily_max = emp.daily_max_hours or 8
                    print(f"[DEBUG] {emp.first_name}: weekly {existing_hours:.1f}+{work_hours:.1f}<={emp.weekly_hours}, daily {existing_hours_today:.1f}+{work_hours:.1f}<={daily_max}", flush=True)

                    # ===== Check for overtime (> 9 hours total in a day) =====
                    daily_total_with_shift = existing_hours_today + total_shift_hours
                    has_overtime = daily_total_with_shift > 9
                    
                    if has_overtime:
                        overtime_warnings.append({
                            'employee_id': emp.id,
                            'employee_name': f"{emp.first_name} {emp.last_name}",
                            'date': current_date.isoformat(),
                            'shift_hours': total_shift_hours,
                            'existing_daily_hours': existing_hours_today,
                            'total_daily_hours': daily_total_with_shift,
                            'total_weekly_hours': existing_hours + work_hours,
                            'message': f"Total {daily_total_with_shift:.1f}h on {current_date} (includes {total_shift_hours}h shift)"
                        })
                        print(f"[DEBUG] ⚠️  OVERTIME: {emp.first_name} would work {daily_total_with_shift:.1f} hours on {current_date}", flush=True)

                    if (existing_hours + work_hours <= emp.weekly_hours and
                        existing_hours_today + work_hours <= daily_max):
                        print(f"[DEBUG] ✓ Creating schedule for {emp.first_name} on {current_date}", flush=True)
                        # Create schedule
                        schedule = Schedule(
                            department_id=department_id,
                            employee_id=emp.id,
                            role_id=shift.role_id,
                            shift_id=shift.id,
                            date=current_date,
                            start_time=shift.start_time,
                            end_time=shift.end_time,
                            status="scheduled"
                        )
                        db.add(schedule)
                        schedules_created += 1
                        assigned_count += 1
                        
                        if assigned_count >= shift.max_emp:
                            break  # Max employees for this shift on this day
                    else:
                        print(f"[DEBUG] ✗ {emp.first_name} failed hours check on {current_date}", flush=True)

                # Ensure minimum employees are assigned
                if assigned_count < shift.min_emp:
                    feedback.append(f"Warning: {shift.name} on {current_date} has {assigned_count} employees (min: {shift.min_emp})")

            current_date += timedelta(days=1)

        await db.commit()

        feedback.insert(0, f"Successfully generated {schedules_created} schedules")
        
        # Add overtime warnings to feedback
        if overtime_warnings:
            feedback.append(f"⚠️  {len(overtime_warnings)} overtime alert(s) - shifts exceed 9 hours on that day")

        return {
            "success": True,
            "schedules_created": schedules_created,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "feedback": feedback,
            "overtime_warnings": overtime_warnings,
            "schedules": []
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in schedule generation: {str(e)}", flush=True)
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Schedule generation error: {str(e)}")


@app.get("/schedules/conflicts")
async def check_schedule_conflicts(
    start_date: date,
    end_date: date,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    """Check for scheduling conflicts"""
    manager_dept = await get_manager_department(current_user, db)
    if not manager_dept:
        raise HTTPException(status_code=400, detail="Manager department not found")
    
    result = await db.execute(
        select(Schedule).filter(
            Schedule.department_id == manager_dept,
            Schedule.date >= start_date,
            Schedule.date <= end_date
        )
    )
    schedules = result.scalars().all()

    # Group by employee and date
    employee_schedules = defaultdict(lambda: defaultdict(list))
    for schedule in schedules:
        employee_schedules[schedule.employee_id][schedule.date].append(schedule)

    conflicts = []
    for emp_id, dates in employee_schedules.items():
        for date, scheds in dates.items():
            if len(scheds) > 1:
                conflicts.append({
                    "employee_id": emp_id,
                    "date": date.isoformat(),
                    "conflicting_schedules": [
                        {
                            "id": s.id,
                            "role_id": s.role_id,
                            "time": f"{s.start_time} - {s.end_time}"
                        }
                        for s in scheds
                    ]
                })

    return {
        "conflicts_found": len(conflicts),
        "conflicts": conflicts
    }


# ==================== ATTENDANCE MANAGEMENT ====================

@app.post("/attendance/record", response_model=AttendanceResponse)
async def record_attendance(
    attendance_data: AttendanceCreate,
    current_user: User = Depends(require_employee),
    db: AsyncSession = Depends(get_db)
):
    """
    Record attendance with check-in time
    Calculates status based on scheduled time
    """
    from datetime import datetime as dt
    
    today = date.today()
    
    # Check if already checked in
    result = await db.execute(
        select(Attendance).filter(
            Attendance.employee_id == current_user.employee_id,
            Attendance.date == today,
            Attendance.in_time.isnot(None)
        )
    )
    if result.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Already checked in today")
    
    # Get schedule for today
    if attendance_data.schedule_id:
        result = await db.execute(
            select(Schedule).filter(
                Schedule.id == attendance_data.schedule_id,
                Schedule.employee_id == current_user.employee_id,
                Schedule.date == today
            )
        )
    else:
        result = await db.execute(
            select(Schedule).filter(
                Schedule.employee_id == current_user.employee_id,
                Schedule.date == today
            )
        )
    
    schedule = result.scalar_one_or_none()
    if not schedule:
        raise HTTPException(status_code=400, detail="No scheduled shift for today")
    
    # Calculate status based on check-in time
    try:
        in_time_parts = attendance_data.in_time.split(':')
        in_hour, in_min = int(in_time_parts[0]), int(in_time_parts[1])
        in_minutes = in_hour * 60 + in_min
        
        start_time_parts = schedule.start_time.split(':')
        start_hour, start_min = int(start_time_parts[0]), int(start_time_parts[1])
        start_minutes = start_hour * 60 + start_min
        
        diff_minutes = in_minutes - start_minutes
        
        if diff_minutes <= 0:
            status_val = "onTime"
        elif diff_minutes <= 15:
            status_val = "slightlyLate"
        elif diff_minutes <= 60:
            status_val = "late"
        else:
            status_val = "veryLate"
    except Exception as e:
        raise HTTPException(status_code=400, detail="Invalid time format")
    
    # Create attendance record
    attendance = Attendance(
        employee_id=current_user.employee_id,
        schedule_id=attendance_data.schedule_id,
        date=today,
        in_time=attendance_data.in_time,
        status=status_val,
        notes=attendance_data.notes
    )
    
    db.add(attendance)
    await db.commit()
    await db.refresh(attendance)
    
    return attendance


@app.put("/attendance/{attendance_id}/checkout")
async def record_checkout(
    attendance_id: int,
    checkout_data: AttendanceUpdate,
    current_user: User = Depends(require_employee),
    db: AsyncSession = Depends(get_db)
):
    """
    Record check-out time and calculate worked hours
    """
    result = await db.execute(
        select(Attendance).filter(
            Attendance.id == attendance_id,
            Attendance.employee_id == current_user.employee_id
        )
    )
    attendance = result.scalar_one_or_none()
    
    if not attendance:
        raise HTTPException(status_code=404, detail="Attendance record not found")
    
    if not checkout_data.out_time:
        raise HTTPException(status_code=400, detail="Out time is required")
    
    # Calculate worked hours
    if attendance.in_time and checkout_data.out_time:
        try:
            in_parts = attendance.in_time.split(':')
            out_parts = checkout_data.out_time.split(':')
            
            in_minutes = int(in_parts[0]) * 60 + int(in_parts[1])
            out_minutes = int(out_parts[0]) * 60 + int(out_parts[1])
            
            if out_minutes < in_minutes:
                out_minutes += 24 * 60  # Handle overnight shifts
            
            total_minutes = out_minutes - in_minutes
            
            # Get role for break time
            schedule = await db.get(Schedule, attendance.schedule_id)
            if schedule:
                role = await db.get(Role, schedule.role_id)
                break_minutes = role.break_minutes if role else 0
            else:
                break_minutes = 0
            
            # Calculate worked hours
            worked_minutes = total_minutes - break_minutes
            worked_hours = max(0, worked_minutes / 60)
            
            attendance.worked_hours = round(worked_hours, 2)
            attendance.break_minutes = break_minutes
            
            # Calculate overtime considering approved overtime
            if schedule:
                emp = await db.get(Employee, attendance.employee_id)
                if emp and worked_hours > emp.daily_max_hours:
                    actual_overtime = worked_hours - emp.daily_max_hours
                    
                    # Get approved overtime for this date if exists
                    approved_overtime_result = await db.execute(
                        select(OvertimeRequest).filter(
                            OvertimeRequest.employee_id == attendance.employee_id,
                            OvertimeRequest.request_date == attendance.date,
                            OvertimeRequest.status == OvertimeStatus.APPROVED
                        )
                    )
                    overtime_request = approved_overtime_result.scalar_one_or_none()
                    
                    if overtime_request:
                        # Use minimum of actual overtime and approved overtime
                        # This ensures we show approved amount if they worked more, or actual if less
                        overtime_hours = min(actual_overtime, overtime_request.request_hours)
                        attendance.overtime_hours = round(overtime_hours, 2)
                    else:
                        # No approved overtime, show actual overtime worked
                        attendance.overtime_hours = round(actual_overtime, 2)
        except Exception as e:
            raise HTTPException(status_code=400, detail="Invalid time format")
    
    attendance.out_time = checkout_data.out_time
    attendance.out_status = checkout_data.out_status
    attendance.notes = checkout_data.notes or attendance.notes
    
    await db.commit()
    await db.refresh(attendance)
    
    return attendance


@app.get("/attendance/weekly/{employee_id}")
async def get_weekly_attendance(
    employee_id: int,
    start_date: date,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Get weekly attendance for an employee"""
    # Check authorization
    if current_user.user_type == UserType.EMPLOYEE and current_user.employee_id != employee_id:
        raise HTTPException(status_code=403, detail="Cannot view other employees' attendance")
    
    # Calculate week end
    end_date = start_date + timedelta(days=6)
    
    # Get attendance records
    result = await db.execute(
        select(Attendance).filter(
            Attendance.employee_id == employee_id,
            Attendance.date >= start_date,
            Attendance.date <= end_date
        ).order_by(Attendance.date)
    )
    records = result.scalars().all()
    
    # Calculate weekly stats
    total_worked = sum(r.worked_hours for r in records)
    total_overtime = sum(r.overtime_hours for r in records)
    on_time_count = len([r for r in records if r.status == "onTime"])
    late_count = len([r for r in records if r.status in ["slightlyLate", "late", "veryLate"]])
    
    return {
        "employee_id": employee_id,
        "week_start": start_date.isoformat(),
        "week_end": end_date.isoformat(),
        "records": [
            {
                "date": r.date.isoformat(),
                "in_time": r.in_time,
                "out_time": r.out_time,
                "status": r.status,
                "worked_hours": r.worked_hours,
                "overtime_hours": r.overtime_hours
            }
            for r in records
        ],
        "summary": {
            "total_days_worked": len(records),
            "total_worked_hours": round(total_worked, 2),
            "total_overtime_hours": round(total_overtime, 2),
            "on_time_count": on_time_count,
            "late_count": late_count,
            "on_time_percentage": round((on_time_count / len(records) * 100) if records else 0, 2)
        }
    }


@app.get("/attendance/summary")
async def get_attendance_summary(
    start_date: date,
    end_date: date,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Get attendance summary for department or individual"""
    query = select(Attendance).filter(
        Attendance.date >= start_date,
        Attendance.date <= end_date
    )
    
    if current_user.user_type == UserType.EMPLOYEE:
        query = query.filter(Attendance.employee_id == current_user.employee_id)
    elif current_user.user_type == UserType.MANAGER:
        manager_dept = await get_manager_department(current_user, db)
        if not manager_dept:
            return []
        # Get employees in manager's department
        subquery = select(Employee.id).filter(
            Employee.department_id == manager_dept
        )
        query = query.filter(Attendance.employee_id.in_(subquery))
    
    result = await db.execute(query)
    records = result.scalars().all()
    
    # Group by employee
    emp_stats = defaultdict(lambda: {
        "total_worked_hours": 0,
        "total_overtime": 0,
        "on_time": 0,
        "late": 0,
        "total_days": 0
    })
    
    for record in records:
        stats = emp_stats[record.employee_id]
        stats["total_worked_hours"] += record.worked_hours
        stats["total_overtime"] += record.overtime_hours
        if record.status == "onTime":
            stats["on_time"] += 1
        elif record.status in ["slightlyLate", "late", "veryLate"]:
            stats["late"] += 1
        stats["total_days"] += 1
    
    # Convert to list with employee details
    summary_list = []
    for emp_id, stats in emp_stats.items():
        emp = await db.get(Employee, emp_id)
        if emp:
            summary_list.append({
                "employee_id": emp_id,
                "employee_name": f"{emp.first_name} {emp.last_name}",
                "total_worked_hours": round(stats["total_worked_hours"], 2),
                "total_overtime": round(stats["total_overtime"], 2),
                "on_time_percentage": round((stats["on_time"] / stats["total_days"] * 100) if stats["total_days"] > 0 else 0, 2),
                "late_count": stats["late"],
                "days_worked": stats["total_days"]
            })
    
    return {
        "period": {
            "start": start_date.isoformat(),
            "end": end_date.isoformat()
        },
        "summary": sorted(summary_list, key=lambda x: x["total_worked_hours"], reverse=True)
    }


# ===== UNAVAILABILITY MANAGEMENT =====

@app.post("/unavailability", response_model=UnavailabilityResponse)
async def create_unavailability(
    unavail: UnavailabilityCreate,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Mark an employee as unavailable for a specific date (manager only)"""
    if current_user.user_type not in [UserType.ADMIN, UserType.MANAGER]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check employee belongs to manager's department
    result = await db.execute(select(Employee).filter(Employee.id == unavail.employee_id))
    employee = result.scalar_one_or_none()
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    if current_user.user_type == UserType.MANAGER:
        manager_dept = await get_manager_department(current_user, db)
        if not manager_dept or employee.department_id != manager_dept:
            raise HTTPException(status_code=404, detail="Employee not found")
    
    from app.models import Unavailability
    unavailability = Unavailability(
        employee_id=unavail.employee_id,
        date=unavail.date,
        reason=unavail.reason
    )
    db.add(unavailability)
    await db.commit()
    await db.refresh(unavailability)
    return unavailability


@app.get("/unavailability", response_model=List[UnavailabilityResponse])
async def list_unavailability(
    employee_id: int = None,
    start_date: date = None,
    end_date: date = None,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """List unavailability records for department (manager) or specific employee (employee)"""
    from app.models import Unavailability
    
    query = select(Unavailability)
    
    if current_user.user_type == UserType.MANAGER:
        manager_dept = await get_manager_department(current_user, db)
        if not manager_dept:
            raise HTTPException(status_code=403, detail="Not authorized")
        # Manager sees unavailability for employees in their department
        if employee_id:
            result = await db.execute(select(Employee).filter(Employee.id == employee_id))
            employee = result.scalar_one_or_none()
            if not employee or employee.department_id != manager_dept:
                raise HTTPException(status_code=404, detail="Employee not found")
            query = query.filter(Unavailability.employee_id == employee_id)
    elif current_user.user_type == UserType.EMPLOYEE:
        # Employee sees only their own unavailability
        result = await db.execute(select(Employee).filter(Employee.user_id == current_user.id))
        employee = result.scalar_one_or_none()
        if not employee:
            raise HTTPException(status_code=404, detail="Employee record not found")
        query = query.filter(Unavailability.employee_id == employee.id)
    
    if start_date:
        query = query.filter(Unavailability.date >= start_date)
    if end_date:
        query = query.filter(Unavailability.date <= end_date)
    
    result = await db.execute(query.order_by(Unavailability.date))
    unavailability_list = result.scalars().all()
    return unavailability_list


@app.delete("/unavailability/{unavailability_id}")
async def delete_unavailability(
    unavailability_id: int,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Delete an unavailability record (manager only)"""
    if current_user.user_type not in [UserType.ADMIN, UserType.MANAGER]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    from app.models import Unavailability
    result = await db.execute(select(Unavailability).filter(Unavailability.id == unavailability_id))
    unavailability = result.scalar_one_or_none()
    
    if not unavailability:
        raise HTTPException(status_code=404, detail="Unavailability record not found")
    
    # Verify employee belongs to manager's department
    result = await db.execute(select(Employee).filter(Employee.id == unavailability.employee_id))
    employee = result.scalar_one_or_none()
    
    if current_user.user_type == UserType.MANAGER:
        manager_dept = await get_manager_department(current_user, db)
        if not manager_dept or employee.department_id != manager_dept:
            raise HTTPException(status_code=403, detail="Not authorized")
    
    await db.delete(unavailability)
    await db.commit()
    
    return {"detail": "Unavailability record deleted successfully"}


# ===== SHIFT MANAGEMENT =====

@app.post("/shifts", response_model=ShiftResponse)
async def create_shift(
    shift: ShiftCreate,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Create a new shift type for a role (manager only)"""
    if current_user.user_type not in [UserType.ADMIN, UserType.MANAGER]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Verify role belongs to manager's department
    result = await db.execute(select(Role).filter(Role.id == shift.role_id))
    role = result.scalar_one_or_none()
    
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    if current_user.user_type == UserType.MANAGER:
        manager_dept = await get_manager_department(current_user, db)
        if not manager_dept or role.department_id != manager_dept:
            raise HTTPException(status_code=404, detail="Role not found")
    
    from app.models import Shift
    new_shift = Shift(
        role_id=shift.role_id,
        name=shift.name,
        start_time=shift.start_time,
        end_time=shift.end_time,
        priority=shift.priority,
        min_emp=shift.min_emp,
        max_emp=shift.max_emp,
        schedule_config=shift.schedule_config
    )
    db.add(new_shift)
    await db.commit()
    await db.refresh(new_shift)
    return new_shift


@app.get("/shifts", response_model=List[ShiftResponse])
async def list_shifts(
    role_id: int = None,
    include_inactive: bool = False,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """List shifts for a role or department (manager)"""
    query = select(Shift)

    if role_id:
        query = query.filter(Shift.role_id == role_id)

    if not include_inactive:
        query = query.filter(Shift.is_active == True)

    if current_user.user_type == UserType.MANAGER:
        manager_dept = await get_manager_department(current_user, db)
        if not manager_dept:
            raise HTTPException(status_code=403, detail="Not authorized")
        query = query.join(Role).filter(Role.department_id == manager_dept)
    elif current_user.user_type != UserType.ADMIN:
        raise HTTPException(status_code=403, detail="Not authorized")

    result = await db.execute(query.order_by(Shift.created_at.desc()))
    shifts = result.scalars().all()
    return shifts


@app.put("/shifts/{shift_id}", response_model=ShiftResponse)
async def update_shift(
    shift_id: int,
    shift_update: ShiftUpdate,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Update a shift (manager only)"""
    if current_user.user_type not in [UserType.ADMIN, UserType.MANAGER]:
        raise HTTPException(status_code=403, detail="Not authorized")

    from app.models import Shift
    result = await db.execute(select(Shift).filter(Shift.id == shift_id))
    shift = result.scalar_one_or_none()

    if not shift:
        raise HTTPException(status_code=404, detail="Shift not found")

    # Verify shift's role belongs to manager's department
    result = await db.execute(select(Role).filter(Role.id == shift.role_id))
    role = result.scalar_one_or_none()

    if current_user.user_type == UserType.MANAGER:
        manager_dept = await get_manager_department(current_user, db)
        if not manager_dept or role.department_id != manager_dept:
            raise HTTPException(status_code=403, detail="Not authorized")

    # Update only provided fields
    for key, value in shift_update.dict(exclude_unset=True).items():
        if value is not None:
            setattr(shift, key, value)

    await db.commit()
    await db.refresh(shift)
    return shift


@app.delete("/shifts/{shift_id}")
async def delete_shift(
    shift_id: int,
    hard_delete: bool = False,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Delete a shift (manager only). Supports soft delete by default with optional permanent delete."""
    if current_user.user_type not in [UserType.ADMIN, UserType.MANAGER]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    result = await db.execute(select(Shift).filter(Shift.id == shift_id))
    shift = result.scalar_one_or_none()
    
    if not shift:
        raise HTTPException(status_code=404, detail="Shift not found")
    
    # Verify shift's role belongs to manager's department
    result = await db.execute(select(Role).filter(Role.id == shift.role_id))
    role = result.scalar_one_or_none()
    
    if current_user.user_type == UserType.MANAGER:
        manager_dept = await get_manager_department(current_user, db)
        if not manager_dept or role.department_id != manager_dept:
            raise HTTPException(status_code=403, detail="Not authorized")
    
    if hard_delete:
        await db.delete(shift)
        await db.commit()
        return {"detail": "Shift permanently deleted"}
    
    if not shift.is_active:
        return {"detail": "Shift already inactive"}
    
    shift.is_active = False
    await db.commit()
    
    return {"detail": "Shift deleted successfully"}


# ===== OVERTIME MANAGEMENT =====

@app.post("/overtime-requests", response_model=OvertimeRequestResponse)
async def create_overtime_request(
    request_data: OvertimeRequestCreate,
    current_user: User = Depends(require_employee),
    db: AsyncSession = Depends(get_db)
):
    """Employee submits an overtime request"""
    # Get employee record
    emp_result = await db.execute(
        select(Employee).filter(Employee.user_id == current_user.id)
    )
    employee = emp_result.scalar_one_or_none()
    
    if not employee:
        raise HTTPException(status_code=400, detail="Employee record not found")
    
    # Create overtime request
    ot_request = OvertimeRequest(
        employee_id=employee.id,
        request_date=request_data.request_date,
        from_time=request_data.from_time,
        to_time=request_data.to_time,
        request_hours=request_data.request_hours,
        reason=request_data.reason,
        status=OvertimeStatus.PENDING
    )
    
    db.add(ot_request)
    await db.commit()
    await db.refresh(ot_request)
    
    return ot_request


@app.get("/overtime-requests", response_model=List[OvertimeRequestResponse])
async def list_overtime_requests(
    status: str = None,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """List overtime requests. Managers see pending requests, employees see their own"""
    query = select(OvertimeRequest)
    
    if current_user.user_type == UserType.EMPLOYEE:
        # Employees see their own requests
        emp_result = await db.execute(
            select(Employee).filter(Employee.user_id == current_user.id)
        )
        employee = emp_result.scalar_one_or_none()
        if not employee:
            raise HTTPException(status_code=400, detail="Employee record not found")
        query = query.filter(OvertimeRequest.employee_id == employee.id)
    elif current_user.user_type == UserType.MANAGER:
        # Managers see pending requests for their department
        manager_dept = await get_manager_department(current_user, db)
        if not manager_dept:
            raise HTTPException(status_code=403, detail="Not authorized")
        
        # Join with Employee to filter by department
        query = query.join(Employee).filter(
            Employee.department_id == manager_dept
        )
        if not status:
            status = OvertimeStatus.PENDING
    else:  # ADMIN
        if not status:
            status = OvertimeStatus.PENDING
    
    if status:
        query = query.filter(OvertimeRequest.status == status)
    
    result = await db.execute(query.order_by(OvertimeRequest.request_date.desc()))
    return result.scalars().all()


@app.put("/overtime-requests/{request_id}/approve", response_model=OvertimeRequestResponse)
async def approve_overtime_request(
    request_id: int,
    approval_data: dict,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    """Manager approves an overtime request"""
    result = await db.execute(
        select(OvertimeRequest).filter(OvertimeRequest.id == request_id)
    )
    ot_request = result.scalar_one_or_none()
    
    if not ot_request:
        raise HTTPException(status_code=404, detail="Overtime request not found")
    
    # Verify manager's authority over employee's department
    emp_result = await db.execute(
        select(Employee).filter(Employee.id == ot_request.employee_id)
    )
    employee = emp_result.scalar_one_or_none()
    
    manager_dept = await get_manager_department(current_user, db)
    if not manager_dept or employee.department_id != manager_dept:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    ot_request.status = OvertimeStatus.APPROVED
    ot_request.approved_at = datetime.utcnow()
    ot_request.approval_notes = approval_data.get("approval_notes", "")
    
    await db.commit()
    await db.refresh(ot_request)
    
    # ========== CREATE NOTIFICATION FOR EMPLOYEE ==========
    from app.helpers import create_notification
    if employee and employee.user_id:
        await create_notification(
            db,
            employee.user_id,
            "Overtime Approved",
            f"Your overtime request for {ot_request.request_date} ({ot_request.request_hours} hours) has been approved.",
            notification_type="overtime_approved",
            related_id=ot_request.id
        )
    
    return ot_request


@app.put("/overtime-requests/{request_id}/reject", response_model=OvertimeRequestResponse)
async def reject_overtime_request(
    request_id: int,
    rejection_data: dict,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    """Manager rejects an overtime request"""
    result = await db.execute(
        select(OvertimeRequest).filter(OvertimeRequest.id == request_id)
    )
    ot_request = result.scalar_one_or_none()
    
    if not ot_request:
        raise HTTPException(status_code=404, detail="Overtime request not found")
    
    # Verify manager's authority over employee's department
    emp_result = await db.execute(
        select(Employee).filter(Employee.id == ot_request.employee_id)
    )
    employee = emp_result.scalar_one_or_none()
    
    manager_dept = await get_manager_department(current_user, db)
    if not manager_dept or employee.department_id != manager_dept:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    ot_request.status = OvertimeStatus.REJECTED
    ot_request.approved_at = datetime.utcnow()
    ot_request.approval_notes = rejection_data.get("approval_notes", "")
    
    await db.commit()
    await db.refresh(ot_request)
    
    # ========== CREATE NOTIFICATION FOR EMPLOYEE ==========
    from app.helpers import create_notification
    if employee and employee.user_id:
        await create_notification(
            db,
            employee.user_id,
            "Overtime Rejected",
            f"Your overtime request for {ot_request.request_date} has been rejected.\nReason: {rejection_data.get('approval_notes', 'No reason provided')}",
            notification_type="overtime_rejected",
            related_id=ot_request.id
        )
    
    return ot_request


@app.post("/manager/overtime-approve", response_model=OvertimeRequestResponse)
async def manager_approve_overtime(
    approve_data: dict,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    """
    Manager directly approves overtime for an employee.
    Creates and approves in one step.
    
    Request body:
    {
        "employee_id": 1,
        "request_date": "2025-12-19",
        "from_time": "18:00",
        "to_time": "19:00",
        "request_hours": 1.0,
        "reason": "Project deadline"
    }
    """
    employee_id = approve_data.get("employee_id")
    request_date = approve_data.get("request_date")
    from_time = approve_data.get("from_time")
    to_time = approve_data.get("to_time")
    request_hours = approve_data.get("request_hours", 0)
    reason = approve_data.get("reason", "Manager approved")
    
    # Verify manager's authority
    emp_result = await db.execute(
        select(Employee).filter(Employee.id == employee_id)
    )
    employee = emp_result.scalar_one_or_none()
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    manager_dept = await get_manager_department(current_user, db)
    if not manager_dept or employee.department_id != manager_dept:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if already approved for this date
    existing = await db.execute(
        select(OvertimeRequest).filter(
            OvertimeRequest.employee_id == employee_id,
            OvertimeRequest.request_date == datetime.strptime(request_date, "%Y-%m-%d").date(),
            OvertimeRequest.status == OvertimeStatus.APPROVED
        )
    )
    
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Overtime already approved for this date")
    
    # Create and immediately approve
    ot_request = OvertimeRequest(
        employee_id=employee_id,
        request_date=datetime.strptime(request_date, "%Y-%m-%d").date(),
        from_time=from_time,
        to_time=to_time,
        request_hours=float(request_hours),
        reason=reason,
        status=OvertimeStatus.APPROVED,
        manager_id=current_user.id,
        approved_at=datetime.utcnow(),
        manager_notes="Approved by manager"
    )
    
    db.add(ot_request)
    await db.commit()
    await db.refresh(ot_request)
    
    # ========== CREATE NOTIFICATION FOR EMPLOYEE ==========
    from app.helpers import create_notification
    if employee and employee.user_id:
        await create_notification(
            db,
            employee.user_id,
            "Overtime Approved",
            f"Your overtime request for {request_date} ({request_hours} hours) from {from_time} to {to_time} has been approved.",
            notification_type="overtime_approved",
            related_id=ot_request.id
        )
    
    return ot_request
    await db.commit()
    await db.refresh(ot_request)
    
    return ot_request


@app.get("/overtime/tracking", response_model=List[OvertimeTrackingResponse])
async def get_overtime_tracking(
    employee_id: int = None,
    year: int = None,
    month: int = None,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Get overtime tracking for employee(s). Returns monthly allocation and usage."""
    if year is None:
        year = date.today().year
    if month is None:
        month = date.today().month
    
    query = select(OvertimeTracking).filter(
        OvertimeTracking.year == year,
        OvertimeTracking.month == month
    )
    
    if current_user.user_type == UserType.EMPLOYEE:
        # Employees see their own tracking
        emp_result = await db.execute(
            select(Employee).filter(Employee.user_id == current_user.id)
        )
        employee = emp_result.scalar_one_or_none()
        if not employee:
            raise HTTPException(status_code=400, detail="Employee record not found")
        query = query.filter(OvertimeTracking.employee_id == employee.id)
    elif current_user.user_type == UserType.MANAGER:
        # Managers see tracking for their department
        manager_dept = await get_manager_department(current_user, db)
        if not manager_dept:
            raise HTTPException(status_code=403, detail="Not authorized")
        
        query = query.join(Employee).filter(
            Employee.department_id == manager_dept
        )
        
        if employee_id:
            query = query.filter(OvertimeTracking.employee_id == employee_id)
    elif employee_id:
        query = query.filter(OvertimeTracking.employee_id == employee_id)
    
    result = await db.execute(query)
    tracking_records = result.scalars().all()
    
    # If no records exist, create them for the month
    if not tracking_records:
        if current_user.user_type == UserType.EMPLOYEE:
            emp_result = await db.execute(
                select(Employee).filter(Employee.user_id == current_user.id)
            )
            employee = emp_result.scalar_one_or_none()
            if employee:
                tracking = OvertimeTracking(
                    employee_id=employee.id,
                    year=year,
                    month=month,
                    allocated_hours=8,  # 8 hours per month baseline
                    used_hours=0.0,
                    remaining_hours=8
                )
                db.add(tracking)
                await db.commit()
                tracking_records = [tracking]
    
    return tracking_records


@app.post("/overtime/check-availability")
async def check_overtime_availability(
    employee_id: int,
    requested_hours: float,
    year: int = None,
    month: int = None,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    """Check if employee has sufficient overtime available for a schedule"""
    if year is None:
        year = date.today().year
    if month is None:
        month = date.today().month
    
    # Verify manager can access this employee
    emp_result = await db.execute(
        select(Employee).filter(Employee.id == employee_id)
    )
    employee = emp_result.scalar_one_or_none()
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    manager_dept = await get_manager_department(current_user, db)
    if not manager_dept or employee.department_id != manager_dept:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get tracking for the month
    result = await db.execute(
        select(OvertimeTracking).filter(
            OvertimeTracking.employee_id == employee_id,
            OvertimeTracking.year == year,
            OvertimeTracking.month == month
        )
    )
    tracking = result.scalar_one_or_none()
    
    if not tracking:
        # Create default tracking if doesn't exist
        tracking = OvertimeTracking(
            employee_id=employee_id,
            year=year,
            month=month,
            allocated_hours=8,
            used_hours=0.0,
            remaining_hours=8
        )
        db.add(tracking)
        await db.commit()
        await db.refresh(tracking)
    
    available = tracking.remaining_hours >= requested_hours
    
    return {
        "employee_id": employee_id,
        "employee_name": f"{employee.first_name} {employee.last_name}",
        "year": year,
        "month": month,
        "allocated_hours": tracking.allocated_hours,
        "used_hours": tracking.used_hours,
        "remaining_hours": tracking.remaining_hours,
        "requested_hours": requested_hours,
        "available": available,
        "message": f"Sufficient overtime available" if available else f"Insufficient overtime. Required: {requested_hours}h, Available: {tracking.remaining_hours}h"
    }


@app.post("/overtime/worked", response_model=OvertimeWorkedResponse)
async def record_overtime_worked(
    worked_data: dict,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    """Record overtime hours worked on a specific day"""
    employee_id = worked_data.get("employee_id")
    work_date = datetime.strptime(worked_data.get("work_date"), "%Y-%m-%d").date()
    overtime_hours = worked_data.get("overtime_hours", 0)
    
    # Verify manager can access this employee
    emp_result = await db.execute(
        select(Employee).filter(Employee.id == employee_id)
    )
    employee = emp_result.scalar_one_or_none()
    
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    manager_dept = await get_manager_department(current_user, db)
    if not manager_dept or employee.department_id != manager_dept:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Create OvertimeWorked record
    ot_worked = OvertimeWorked(
        employee_id=employee_id,
        work_date=work_date,
        overtime_hours=overtime_hours
    )
    
    db.add(ot_worked)
    
    # Update OvertimeTracking for that month
    month = work_date.month
    year = work_date.year
    
    result = await db.execute(
        select(OvertimeTracking).filter(
            OvertimeTracking.employee_id == employee_id,
            OvertimeTracking.year == year,
            OvertimeTracking.month == month
        )
    )
    tracking = result.scalar_one_or_none()
    
    if not tracking:
        # Create if doesn't exist
        tracking = OvertimeTracking(
            employee_id=employee_id,
            year=year,
            month=month,
            allocated_hours=8,
            used_hours=overtime_hours,
            remaining_hours=8 - overtime_hours
        )
        db.add(tracking)
    else:
        # Update existing
        tracking.used_hours += overtime_hours
        tracking.remaining_hours = tracking.allocated_hours - tracking.used_hours
    
    await db.commit()
    await db.refresh(ot_worked)
    
    return ot_worked


@app.get("/overtime/worked", response_model=List[OvertimeWorkedResponse])
async def list_overtime_worked(
    employee_id: int = None,
    start_date: date = None,
    end_date: date = None,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """List overtime worked records"""
    query = select(OvertimeWorked)
    
    if current_user.user_type == UserType.EMPLOYEE:
        # Employees see their own overtime
        emp_result = await db.execute(
            select(Employee).filter(Employee.user_id == current_user.id)
        )
        employee = emp_result.scalar_one_or_none()
        if not employee:
            raise HTTPException(status_code=400, detail="Employee record not found")
        query = query.filter(OvertimeWorked.employee_id == employee.id)
    elif current_user.user_type == UserType.MANAGER:
        # Managers see overtime for their department
        manager_dept = await get_manager_department(current_user, db)
        if not manager_dept:
            raise HTTPException(status_code=403, detail="Not authorized")
        
        query = query.join(Employee).filter(
            Employee.department_id == manager_dept
        )
        
        if employee_id:
            query = query.filter(OvertimeWorked.employee_id == employee_id)
    elif employee_id:
        query = query.filter(OvertimeWorked.employee_id == employee_id)
    
    if start_date:
        query = query.filter(OvertimeWorked.work_date >= start_date)
    if end_date:
        query = query.filter(OvertimeWorked.work_date <= end_date)
    
    result = await db.execute(query.order_by(OvertimeWorked.work_date.desc()))
    return result.scalars().all()


@app.delete("/admin/roles/all")
async def delete_all_roles(
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Delete all roles (admin only) - for testing/cleanup"""
    if current_user.user_type != UserType.ADMIN:
        raise HTTPException(status_code=403, detail="Only admins can perform this action")

    # Mark all roles as inactive
    result = await db.execute(select(Role).filter(Role.is_active == True))
    roles = result.scalars().all()

    count = 0
    for role in roles:
        role.is_active = False
        count += 1

    await db.commit()

    return {"message": f"Deleted {count} roles", "count": count}


# ==================== COMPREHENSIVE ATTENDANCE MANAGEMENT ====================

@app.get("/attendance/summary-detailed/{employee_id}")
async def get_detailed_attendance_summary(
    employee_id: int,
    period_type: str = 'monthly',
    period_date: Optional[str] = None,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Get comprehensive detailed attendance summary for an employee"""
    
    # Parse period date
    if period_date:
        try:
            period_date = datetime.strptime(period_date, '%Y-%m-%d').date()
        except ValueError:
            period_date = date.today()
    else:
        period_date = date.today()
    
    # Get summary from service
    summary = await AttendanceService.aggregate_attendance_summary(
        db, employee_id, period_type, period_date
    )
    
    return {
        "success": True,
        "data": summary
    }


@app.get("/attendance/comprehensive-report")
async def get_comprehensive_attendance_report(
    start_date: str,
    end_date: str,
    department_id: Optional[int] = None,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    """Get comprehensive attendance report for department or all employees"""
    
    try:
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    # Get employees
    if department_id:
        emp_result = await db.execute(
            select(Employee).where(
                and_(
                    Employee.department_id == department_id,
                    Employee.is_active == True
                )
            )
        )
    else:
        emp_result = await db.execute(select(Employee).where(Employee.is_active == True))
    
    employees = emp_result.scalars().all()
    
    reports = []
    for emp in employees:
        summary = await AttendanceService.aggregate_attendance_summary(
            db, emp.id, 'monthly', start
        )
        reports.append({
            "employee_id": emp.employee_id,
            "employee_name": f"{emp.first_name} {emp.last_name}",
            **summary
        })
    
    return {
        "success": True,
        "period": {
            "start": start.isoformat(),
            "end": end.isoformat()
        },
        "total_employees": len(reports),
        "data": reports
    }


@app.get("/attendance/validate/{employee_id}")
async def validate_attendance_data(
    employee_id: int,
    start_date: str,
    end_date: str,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    """Validate attendance data for completeness and accuracy"""
    
    try:
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    validation_result = await AttendanceService.validate_attendance_data(
        db, employee_id, start, end
    )
    
    return {
        "success": True,
        "data": validation_result
    }


@app.post("/attendance/summary/create")
async def create_attendance_summary(
    employee_id: int,
    period_type: str = 'monthly',
    period_date: Optional[str] = None,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    """Create or update attendance summary record"""
    
    if period_date:
        try:
            period_date = datetime.strptime(period_date, '%Y-%m-%d').date()
        except ValueError:
            period_date = date.today()
    else:
        period_date = date.today()
    
    summary = await AttendanceService.create_or_update_attendance_summary(
        db, employee_id, period_type, period_date
    )
    
    return {
        "success": True,
        "message": "Attendance summary created/updated successfully",
        "data": {
            "id": summary.id,
            "employee_id": summary.employee_id,
            "period_type": summary.period_type,
            "period_date": summary.period_date.isoformat(),
            "total_worked_hours": summary.total_worked_hours,
            "overtime_hours": summary.overtime_hours,
            "night_hours": summary.night_hours
        }
    }


# ==================== PAID LEAVE REMINDER MANAGEMENT ====================

@app.get("/leave/balance-summary/{employee_id}")
async def get_leave_balance_summary(
    employee_id: int,
    year: Optional[int] = None,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Get paid leave balance summary for an employee"""
    
    if year is None:
        year = date.today().year
    
    balance_info = await LeaveReminderService.check_leave_balance(db, employee_id, year)
    
    return {
        "success": True,
        "data": balance_info
    }


@app.get("/leave/trends/{employee_id}")
async def get_leave_trends(
    employee_id: int,
    num_years: int = 3,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Analyze leave trends for an employee"""
    
    trends = await LeaveReminderService.get_leave_trends(db, employee_id, num_years)
    
    return {
        "success": True,
        "data": trends
    }


@app.post("/leave/send-reminders")
async def send_leave_reminders(
    reminder_type: str = 'low_balance',
    threshold_days: int = 3,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    """Send paid leave reminders to employees"""
    
    if reminder_type == 'low_balance':
        reminders = await LeaveReminderService.send_reminders_to_low_balance(db, threshold_days)
    elif reminder_type == 'mid_year':
        reminders = await LeaveReminderService.send_mid_year_reminder(db)
    elif reminder_type == 'year_end':
        reminders = await LeaveReminderService.send_year_end_reminder(db)
    else:
        raise HTTPException(status_code=400, detail="Invalid reminder type")
    
    return {
        "success": True,
        "message": f"Sent {len(reminders)} reminders",
        "reminders_sent": reminders
    }


@app.post("/leave/acknowledge-reminder/{reminder_id}")
async def acknowledge_reminder(
    reminder_id: int,
    action_taken: Optional[str] = None,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Acknowledge a leave reminder"""
    
    reminder = await LeaveReminderService.track_reminder_sent(
        db, reminder_id, action_taken, is_acknowledged=True
    )
    
    if not reminder:
        raise HTTPException(status_code=404, detail="Reminder not found")
    
    return {
        "success": True,
        "message": "Reminder acknowledged",
        "data": {
            "id": reminder.id,
            "employee_id": reminder.employee_id,
            "is_acknowledged": reminder.is_acknowledged,
            "acknowledgment_date": reminder.acknowledgment_date.isoformat() if reminder.acknowledgment_date else None
        }
    }


@app.get("/leave/department-summary/{department_id}")
async def get_department_leave_summary(
    department_id: int,
    year: Optional[int] = None,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    """Get leave summary for all employees in a department"""
    
    if year is None:
        year = date.today().year
    
    summary = await LeaveReminderService.get_department_leave_summary(db, department_id, year)
    
    return {
        "success": True,
        "data": summary
    }


# ==================== PAYROLL WAGE CALCULATION ====================

@app.post("/payroll/configure-employee")
async def configure_employee_wage(
    employee_id: int,
    hourly_rate: float,
    overtime_multiplier: float = 1.5,
    night_shift_multiplier: float = 1.5,
    current_user: User = Depends(require_manager),
    db: AsyncSession = Depends(get_db)
):
    """Configure wage settings for an employee"""
    
    config = await WageCalculationService.get_or_create_wage_config(
        db, employee_id, hourly_rate, overtime_multiplier, night_shift_multiplier
    )
    
    return {
        "success": True,
        "message": "Wage configuration saved",
        "data": {
            "id": config.id,
            "employee_id": config.employee_id,
            "hourly_rate": config.hourly_rate,
            "overtime_multiplier": config.overtime_multiplier,
            "night_shift_multiplier": config.night_shift_multiplier
        }
    }


@app.post("/payroll/process-cycle")
async def process_payroll_cycle(
    start_date: str,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """Process 15-day payroll closing cycle"""
    
    try:
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    # Get or create payroll cycle
    cycle = await WageCalculationService.get_payroll_cycle(db, start)
    
    # Get all employees
    emp_result = await db.execute(
        select(Employee).where(
            and_(
                Employee.is_active == True,
                Employee.wage_config != None
            )
        ).options(selectinload(Employee.wage_config))
    )
    employees = emp_result.scalars().all()
    
    processed = 0
    errors = []
    
    for emp in employees:
        try:
            wage_calc = await WageCalculationService.calculate_wage_for_period(
                db, emp.id, cycle
            )
            processed += 1
        except Exception as e:
            errors.append({
                "employee_id": emp.employee_id,
                "error": str(e)
            })
    
    return {
        "success": True,
        "message": f"Processed {processed} employees",
        "data": {
            "cycle_id": cycle.id,
            "cycle_number": cycle.cycle_number,
            "start_date": cycle.start_date.isoformat(),
            "end_date": cycle.end_date.isoformat(),
            "processed_count": processed,
            "errors": errors
        }
    }


@app.post("/payroll/close-cycle/{cycle_id}")
async def close_payroll_cycle(
    cycle_id: int,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """Close and verify 15-day payroll cycle"""
    
    cycle_result = await db.execute(
        select(PayrollCycle).where(PayrollCycle.id == cycle_id)
    )
    cycle = cycle_result.scalar_one_or_none()
    
    if not cycle:
        raise HTTPException(status_code=404, detail="Cycle not found")
    
    result = await WageCalculationService.verify_and_close_cycle(db, cycle)
    
    return {
        "success": True,
        "message": "Payroll cycle closed",
        "data": result
    }


@app.post("/payroll/confirm-wages/{cycle_id}")
async def confirm_wages(
    cycle_id: int,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """Confirm 18-day wage cycle"""
    
    cycle_result = await db.execute(
        select(PayrollCycle).where(PayrollCycle.id == cycle_id)
    )
    cycle = cycle_result.scalar_one_or_none()
    
    if not cycle:
        raise HTTPException(status_code=404, detail="Cycle not found")
    
    result = await WageCalculationService.confirm_wages(db, cycle)
    
    return {
        "success": True,
        "message": "Wages confirmed",
        "data": result
    }


@app.get("/payroll/wage-summary/{employee_id}")
async def get_wage_summary(
    employee_id: int,
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Get wage summary for an employee over a date range"""
    
    try:
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    summary = await WageCalculationService.get_wage_summary_for_employee(
        db, employee_id, start, end
    )
    
    return {
        "success": True,
        "data": summary
    }


@app.get("/payroll/employee-wages/{employee_id}")
async def export_employee_wages(
    employee_id: int,
    start_date: str,
    end_date: str,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    """Export employee wage data as detailed report"""
    
    try:
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    summary = await WageCalculationService.get_wage_summary_for_employee(
        db, employee_id, start, end
    )
    
    return {
        "success": True,
        "data": summary
    }


@app.get("/payroll/cycles")
async def get_payroll_cycles(
    year: Optional[int] = None,
    is_closed: Optional[bool] = None,
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db)
):
    """Get list of payroll cycles"""
    
    if year is None:
        year = date.today().year
    
    query = select(PayrollCycle).where(PayrollCycle.year == year)
    
    if is_closed is not None:
        query = query.where(PayrollCycle.is_closed == is_closed)
    
    result = await db.execute(query)
    cycles = result.scalars().all()
    
    return {
        "success": True,
        "data": [
            {
                "id": c.id,
                "cycle_number": c.cycle_number,
                "year": c.year,
                "start_date": c.start_date.isoformat(),
                "end_date": c.end_date.isoformat(),
                "closing_date": c.closing_date.isoformat() if c.closing_date else None,
                "confirmation_date": c.confirmation_date.isoformat() if c.confirmation_date else None,
                "is_closed": c.is_closed,
                "is_confirmed": c.is_confirmed,
                "processed_employees": c.processed_employees,
                "confirmed_employees": c.confirmed_employees
            }
            for c in cycles
        ]
    }


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
